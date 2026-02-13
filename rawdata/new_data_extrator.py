from PyQt5.QtGui import QFont, QIcon, QStandardItemModel, QStandardItem, QTextBlock, QTextCursor
from PyQt5.QtCore import Qt, QThread, QObject, pyqtSignal, QCoreApplication
from openpyxl.utils.exceptions import InvalidFileException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoAlertPresentException, NoSuchElementException, StaleElementReferenceException, InvalidSessionIdException, WebDriverException, SessionNotCreatedException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from gspread.utils import rowcol_to_a1
from gspread_formatting import *
from gspread.exceptions import APIError
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from datetime import datetime, date, timedelta
from PyQt5.QtWidgets import *
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.edge.service import Service
import pandas as pd
import chromedriver_autoinstaller
import datetime
import datetime as dt
import numpy as np
import functools
import threading
import openpyxl
import gspread
import json
import time
import glob
import csv
import sys
import os
import re
import shutil
import zipfile
import chardet


class Rawdata_extractor(QWidget):

    def __init__(self):
        super().__init__()
        self.UI초기화()

    def UI초기화(self):

        self.setWindowTitle("Raw data 자동 추출기")
        self.setFixedSize(1100, 900)

# 전체선택 체크박스
        self.all_checkbox = QCheckBox("전체선택/해제", self)
        self.all_checkbox.stateChanged.connect(self.toggle_all_checkboxes)
        self.all_checkbox.move(50,40)

# 매출 group box
        self.sales_group_box = QGroupBox("매출",self)
        self.sales_group_box.setFont(QFont('Helvetia', 20, QFont.Bold))
        self.sales_group_box.move(40,70)
        self.sales_group_box.setFixedSize(400, 250)

    # 카페24
        self.salesCafe24 = QLabel("카페24",self)
        self.salesCafe24.move(50,110)
        self.salesCafe24.setFont(QFont('Helvetia', 14, QFont.Bold))

        # 천명연구소
        self.CMlabs_salesCafe24 = QCheckBox("천명연구소",self)
        self.CMlabs_salesCafe24.move(100,135)
        self.CMlabs_salesCafe24.setFont(QFont('Helvetia', 11))
        
        # 노마셀
        self.know_salesCafe24 = QCheckBox("노마셀",self)
        self.know_salesCafe24.move(210,135)
        self.know_salesCafe24.setFont(QFont('Helvetia', 11))

        # 제니크
        self.zq_salesCafe24 = QCheckBox("제니크",self)
        self.zq_salesCafe24.move(290,135)
        self.zq_salesCafe24.setFont(QFont('Helvetia', 11))

    # 쿠팡
        self.salesCoup = QLabel("쿠팡",self)
        self.salesCoup.move(50,170)
        self.salesCoup.setFont(QFont('Helvetia', 15, QFont.Bold))

        # 노마셀
        self.know_salesCoup = QCheckBox("노마셀",self)
        self.know_salesCoup.move(210,195)
        self.know_salesCoup.setFont(QFont('Helvetia', 11))

    # 네이버
        self.salesNaver = QLabel("네이버",self)
        self.salesNaver.move(50,230)
        self.salesNaver.setFont(QFont('Helvetia', 15, QFont.Bold))

        # 노마셀
        self.know_salesNaver = QCheckBox("노마셀",self)
        self.know_salesNaver.move(210,255)
        self.know_salesNaver.setFont(QFont('Helvetia', 11))


# 광고 group box
        self.advt_group_box = QGroupBox("광고",self)
        self.advt_group_box.setFont(QFont('Helvetia', 20, QFont.Bold))
        self.advt_group_box.move(490,70)
        self.advt_group_box.setFixedSize(400, 460)

    # 쿠팡
        self.advtCoup = QLabel("쿠팡",self)
        self.advtCoup.move(500,110)
        self.advtCoup.setFont(QFont('Helvetia', 15, QFont.Bold))

        # 노마셀
        self.know_advtCoup = QCheckBox("노마셀",self)
        self.know_advtCoup.move(660,135)
        self.know_advtCoup.setFont(QFont('Helvetia', 11))

    # 네이버
        self.advtNaver = QLabel("네이버",self)
        self.advtNaver.move(500,170)
        self.advtNaver.setFont(QFont('Helvetia', 15, QFont.Bold))

        # 노마셀
        self.know_advtNaver = QCheckBox("노마셀",self)
        self.know_advtNaver.move(660,195)
        self.know_advtNaver.setFont(QFont('Helvetia', 11))

        # 제니크
        self.zq_advtNaver = QCheckBox("제니크",self)
        self.zq_advtNaver.move(740,195)
        self.zq_advtNaver.setFont(QFont('Helvetia', 11))

    # GFA
        self.advtGFA = QLabel("GFA",self)
        self.advtGFA.move(500,230)
        self.advtGFA.setFont(QFont('Helvetia', 15, QFont.Bold))

        # 노마셀
        self.know_advtGFA = QCheckBox("노마셀",self)
        self.know_advtGFA.move(660,255)
        self.know_advtGFA.setFont(QFont('Helvetia', 11))

        # 제니크
        self.zq_advtGFA = QCheckBox("제니크",self)
        self.zq_advtGFA.move(740,255)
        self.zq_advtGFA.setFont(QFont('Helvetia', 11))

    # 파워컨텐츠
        self.advtPC = QLabel("파워컨텐츠",self)
        self.advtPC.move(500,290)
        self.advtPC.setFont(QFont('Helvetia', 15, QFont.Bold))

        # 노마셀
        self.know_advtPC = QCheckBox("노마셀",self)
        self.know_advtPC.move(660,315)
        self.know_advtPC.setFont(QFont('Helvetia', 11))

    # 구글
        self.advtGgle = QLabel("구글",self)
        self.advtGgle.move(500,350)
        self.advtGgle.setFont(QFont('Helvetia', 15, QFont.Bold))

        # 노마셀
        self.know_advtGgle = QCheckBox("노마셀",self)
        self.know_advtGgle.move(660,375)
        self.know_advtGgle.setFont(QFont('Helvetia', 11))

    # 메타
        self.advtMeta = QLabel("메타",self)
        self.advtMeta.move(500,410)
        self.advtMeta.setFont(QFont('Helvetia', 15, QFont.Bold))

        # 천명연구소
        self.CMlabs_advtMeta = QCheckBox("천명연구소",self)
        self.CMlabs_advtMeta.move(550,435)
        self.CMlabs_advtMeta.setFont(QFont('Helvetia', 11))

        # 노마셀
        self.know_advtMeta = QCheckBox("노마셀",self)
        self.know_advtMeta.move(660,435)
        self.know_advtMeta.setFont(QFont('Helvetia', 11))

        # 제니크
        self.zq_advtMeta = QCheckBox("제니크",self)
        self.zq_advtMeta.move(740,435)
        self.zq_advtMeta.setFont(QFont('Helvetia', 11))

    # 틱톡
        self.advtTiktok = QLabel("틱톡",self)
        self.advtTiktok.move(500,470)
        self.advtTiktok.setFont(QFont('Helvetia', 15, QFont.Bold))

        # 천명연구소
        self.CMlabs_advtTiktok = QCheckBox("천명연구소",self)
        self.CMlabs_advtTiktok.move(550,495)
        self.CMlabs_advtTiktok.setFont(QFont('Helvetia', 11))

        # 노마셀
        self.know_advtTiktok = QCheckBox("노마셀",self)
        self.know_advtTiktok.move(660,495)
        self.know_advtTiktok.setFont(QFont('Helvetia', 11))

        # 제니크
        self.zq_advtTiktok = QCheckBox("제니크",self)
        self.zq_advtTiktok.move(740,495)
        self.zq_advtTiktok.setFont(QFont('Helvetia', 11))


# 기타 group box
        self.etc_group_box = QGroupBox("기타",self)
        self.etc_group_box.setFont(QFont('Helvetia', 20, QFont.Bold))
        self.etc_group_box.move(40,340)
        self.etc_group_box.setFixedSize(400, 170)

    # 카페24 방문자수
        self.visitors = QLabel("방문자수",self)
        self.visitors.move(50,380)
        self.visitors.setFont(QFont('Helvetia', 15, QFont.Bold))

        # 노마셀
        self.know_visitors = QCheckBox("노마셀",self)
        self.know_visitors.move(210,405)
        self.know_visitors.setFont(QFont('Helvetia', 11))

        # 제니크
        self.zq_visitors = QCheckBox("제니크",self)
        self.zq_visitors.move(290,405)
        self.zq_visitors.setFont(QFont('Helvetia', 11))

    # 카페24 신규가입자수
        self.newMemb = QLabel("신규가입자수",self)
        self.newMemb.move(50,440)
        self.newMemb.setFont(QFont('Helvetia', 15, QFont.Bold))

        # 노마셀
        self.know_newMemb = QCheckBox("노마셀",self)
        self.know_newMemb.move(210,465)
        self.know_newMemb.setFont(QFont('Helvetia', 11))

        # 제니크
        self.zq_newMemb = QCheckBox("제니크",self)
        self.zq_newMemb.move(290,465)
        self.zq_newMemb.setFont(QFont('Helvetia', 11))

        #불러오기 체크박스설정
        self.loadCheckboxState()


# 로그박스
        self.logBox = QTextEdit(self)
        self.logBox.setGeometry(600, 560, 400, 300)
        self.logBox.setReadOnly(True)


# 버튼

    # 다운로드
        # 다운로드폴더 버튼
        self.slt_folder = QPushButton('다운로드폴더',self)
        self.slt_folder.setGeometry(330,571,100,29)
        self.slt_folder.clicked.connect(self.folderopen)
        self.slt_folder.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border-radius: 1.5px;
                border-width: 1px;
                border-color: black;
                border-style: solid;
            }
            QPushButton:hover {
                background-color: rgb(120,120,120);
            }
            QPushButton:pressed {
                background-color: rgb(50, 50, 50);
            }
            """
        )

        # 다운로드폴더 설정저장 버튼
        self.saveButton = QPushButton('설정저장', self)
        self.saveButton.setGeometry(440,571,100,29)
        self.saveButton.clicked.connect(self.saveText)
        self.saveButton.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border-radius: 1.5px;
                border-width: 1px;
                border-color: black;
                border-style: solid;
            }
            QPushButton:hover {
                background-color: rgb(120,120,120);
            }
            QPushButton:pressed {
                background-color: rgb(50, 50, 50);
            }
            """
        )

         # 다운로드폴더 경로
        self.path_folder = QLineEdit(self)
        self.path_folder.setGeometry(80,571,240,27)
        self.path_folder.setStyleSheet(
                        "background-color: white;"
                        "border-radius: 1.5px;"
                        "border-width: 1px;"
                        "border-color: black;"
                        "border-style: solid;")  # 테두리 스타일 추가
        self.path_folder.setReadOnly(True)

    # 크롬폴더
        # 크롬폴더 버튼
        self.chrome_slt_folder = QPushButton('크롬 폴더',self)
        self.chrome_slt_folder.setGeometry(330,620,100,29)
        self.chrome_slt_folder.clicked.connect(self.chromefolderopen)
        self.chrome_slt_folder.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border-radius: 1.5px;
                border-width: 1px;
                border-color: black;
                border-style: solid;
            }
            QPushButton:hover {
                background-color: rgb(120,120,120);
            }
            QPushButton:pressed {
                background-color: rgb(50, 50, 50);
            }
            """
        )

        # 크롬폴더 경로
        self.chrome_path_folder = QLineEdit(self)
        self.chrome_path_folder.setGeometry(80,620,240,27)
        self.chrome_path_folder.setStyleSheet(
                        "background-color: white;"
                        "border-radius: 1.5px;"
                        "border-width: 1px;"
                        "border-color: black;"
                        "border-style: solid;")  # 테두리 스타일 추가
        self.chrome_path_folder.setReadOnly(True)

    # 엣지폴더
        # 엣지폴더 버튼
        self.edge_slt_folder = QPushButton('엣지 폴더',self)
        self.edge_slt_folder.setGeometry(330,670,100,29)
        self.edge_slt_folder.clicked.connect(self.edgefolderopen)
        self.edge_slt_folder.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border-radius: 1.5px;
                border-width: 1px;
                border-color: black;
                border-style: solid;
            }
            QPushButton:hover {
                background-color: rgb(120,120,120);
            }
            QPushButton:pressed {
                background-color: rgb(50, 50, 50);
            }
            """
        )

        # 엣지폴더 경로
        self.edge_path_folder = QLineEdit(self)
        self.edge_path_folder.setGeometry(80,670,240,27)
        self.edge_path_folder.setStyleSheet(
                        "background-color: white;"
                        "border-radius: 1.5px;"
                        "border-width: 1px;"
                        "border-color: black;"
                        "border-style: solid;")  # 테두리 스타일 추가
        self.edge_path_folder.setReadOnly(True)
        self.loadText()  

    # 날짜
        # 날짜 선택
        self.combo = QComboBox(self)
        self.combo.setGeometry(75, 720, 50, 39)
        self.combo.addItems(["1", "2", "3", "4", "5", "6", "7"])
        self.combo.setFont(QFont('Helvetia', 12, QFont.Bold))

        # 날짜 레이블
        self.daybefore = QLabel("일 전까지", self)
        self.daybefore.move(75, 765)
        self.daybefore.setFont(QFont('Helvetia', 12, QFont.Bold))

    # 추출
        # 추출하기
        self.extr_button = QPushButton('추출하기',self)
        self.extr_button.setGeometry(130,720,410,40)
        self.extr_button.clicked.connect(self.extract)
        self.extr_button.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border-radius: 1.5px;
                border-width: 1px;
                border-color: black;
                border-style: solid;
            }
            QPushButton:hover {
                background-color: rgb(120,120,120);
            }
            QPushButton:pressed {
                background-color: rgb(50, 50, 50);
            }
            """
        )




    def toggle_all_checkboxes(self, state):
        self.CMlabs_salesCafe24.setChecked(state == 2)
        self.know_salesCafe24.setChecked(state == 2)
        self.zq_salesCafe24.setChecked(state == 2)

        self.know_salesCoup.setChecked(state == 2)

        self.know_salesNaver.setChecked(state == 2)

        self.know_advtCoup.setChecked(state == 2)

        self.know_advtNaver.setChecked(state == 2)
        self.zq_advtNaver.setChecked(state == 2)

        self.know_advtGFA.setChecked(state == 2)
        self.zq_advtGFA.setChecked(state == 2)

        self.know_advtPC.setChecked(state == 2)

        self.know_advtGgle.setChecked(state == 2)

        self.CMlabs_advtMeta.setChecked(state == 2)
        self.know_advtMeta.setChecked(state == 2)
        self.zq_advtMeta.setChecked(state == 2)

        self.CMlabs_advtTiktok.setChecked(state == 2)
        self.know_advtTiktok.setChecked(state == 2)
        self.zq_advtTiktok.setChecked(state == 2)

        self.know_visitors.setChecked(state == 2)
        self.zq_visitors.setChecked(state == 2)

        self.know_newMemb.setChecked(state == 2)
        self.zq_newMemb.setChecked(state == 2)

    def extract(self):

        # 타겟날짜 변수 저장
        target_days_input = int(self.combo.currentText())

        global download_folder
        download_folder = self.path_folder.text()


        def count_files(folder):
            """ 폴더 내 파일의 개수를 반환합니다. """
            return len([name for name in os.listdir(folder) if os.path.isfile(os.path.join(folder, name))])

        def get_latest_file(folder):
            """ 폴더 내에서 가장 최신의 파일을 반환합니다. """
            files = [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
            latest_file = max(files, key=os.path.getctime)
            return latest_file
        
        def get_previous_latest_file(folder):
            """폴더 내에서 가장 최신 파일을 제외한 이전 파일을 반환합니다."""
            # 폴더 내의 파일들의 전체 경로와 함께 리스트를 생성합니다.
            files = [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
            
            # 파일이 없다면 None을 반환
            if not files:
                return None
            
            # 파일들을 생성 시간 기준으로 정렬합니다.
            files.sort(key=os.path.getctime)
            
            # 가장 최신 파일을 제외한 가장 최신 파일을 찾습니다.
            # 파일이 하나만 있는 경우에는 그 파일이 최신 파일이므로, None을 반환합니다.
            if len(files) > 1:
                previous_latest_file = files[-2]  # 뒤에서 두 번째 항목 선택
                print(previous_latest_file)
                return previous_latest_file
            else:
                return None
            
        def get_nth_latest_file(folder, n):
            """폴더 내에서 n번째로 최신 파일을 반환합니다. n이 1이면 가장 최신, 2면 두 번째로 최신 파일을 반환합니다."""
            # 폴더 내의 파일들의 전체 경로와 함께 리스트를 생성합니다.
            files = [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
            
            # 파일이 없다면 None을 반환
            if not files:
                return None
            
            # 파일들을 생성 시간 기준으로 정렬합니다.
            files.sort(key=os.path.getctime, reverse=True)
            
            # 요청한 순위의 파일을 반환합니다. n이 파일 수보다 많거나 0 이하인 경우 None을 반환합니다.
            if 1 <= n <= len(files):
                nth_latest_file = files[n-1]  # n번째 파일 선택
                print(nth_latest_file)
                return nth_latest_file
            else:
                return None

        def check_download():
            # 다운로드 전의 파일 개수 확인
                initial_file_count = count_files(download_folder)

                # 다운로드 시작 ...

                # 새 파일이 다운로드될 때까지 기다림
                global check
                check = 0
                i = 0

                while i < 20:
                    current_file_count = count_files(download_folder)
                    if current_file_count > initial_file_count:
                        print("A new file has been downloaded.")
                        latest_file = get_latest_file(download_folder)
                        print(f"Downloaded file: {latest_file}")
                        # 여기서 필요한 작업을 수행하세요, 예를 들면 파일 열기 등
                        check = 1
                        break
                    else:
                        print("Still no new file")
                    time.sleep(0.3)  # 폴더 상태를 0.3초마다 체크
                    i += 1
                return check

        def check_data_in_second_row(file_path):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            second_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
            if second_row and any(cell is not None for cell in second_row[0]):
                return True
            return False
        
        def convert_data(data):
            result = []
            for item in data:
                if isinstance(item, str) and '%' in item:
                    result.append(float(item.strip('%')) / 100)
                elif isinstance(item, str) and ',' in item:
                    result.append(int(item.replace(',', '')))
                elif item.isdigit():
                    result.append(int(item))
                else:
                    result.append(item)
            return result
    
        # 날짜 변수
        dayx = datetime.timedelta(days=target_days_input)
        day1 = datetime.timedelta(days=1)
        today = date.today()

        today_date = today.strftime("%d")
        today_month = str(int(today.strftime("%m")))

        weekday_korean = {
            0: '월',
            1: '화',
            2: '수',
            3: '목',
            4: '금',
            5: '토',
            6: '일'
        }

        # 오늘 날짜 구하기
        global today_yday
        global today_tday
        today_yday = today-day1
        today_tday = today-dayx
        today_Tday년월 = (today-dayx).strftime("%Y년 %m월")
        today_Yday년월 = (today-day1).strftime("%Y년 %m월")
        Tday_month월 = str(int(today_tday.strftime("%m"))) + "월"
        Yday_month월 = str(int(today_yday.strftime("%m"))) + "월"
        today_Tday일 = str(int((today-dayx).strftime("%d")))
        today_Yday일 = str(int((today-day1).strftime("%d")))
        today_tday_str = (today-dayx).strftime('%Y-%m-%d')


        weekday_num = today.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
        weekday_numy = today_yday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
        weekday_numt = today_tday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
        # 요일을 한국어로 변환
        weekday_kr = weekday_korean[weekday_num]
        weekday_kry = weekday_korean[weekday_numy]
        weekday_krt = weekday_korean[weekday_numt]

        weekday = f"{today}({weekday_kr})"
        weekday_y = f"{today_yday}({weekday_kry})"
        weekday_t = f"{today_tday}({weekday_krt})"

# 카페24 매출
        def cafe24(url_cafe24, url_cafe24_req, cafe24_id, cafe24_pw, sheet_urlR, sheet_nameR, brand):
            driver = None
            try:
                
                # 크롬 On
                ### chromedriver_autoinstaller.install() 사용 추가
                chromedriver_path = chromedriver_autoinstaller.install()
                chrome_options = webdriver.ChromeOptions()
                chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
                chrome_options.add_argument("--start-maximized") #최대 크기로 시작
                # chrome_options.add_argument('--incognito')
                # chrome_options.add_argument('--window-size=1920,1080')  
                # chrome_options.add_argument('--headless')
                chrome_options.add_experimental_option('detach', True)

                user_data = self.chrome_path_folder.text()
                chrome_options.add_argument(f"user-data-dir={user_data}")
                chrome_options.add_argument("--profile-directory=Profile 1")
                
                user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
                headers = {'user-agent' : user_agent}

                driver = webdriver.Chrome(
                    service=Service(chromedriver_path),
                    options=chrome_options
                )

                
                
                driver.get(url_cafe24)

                ##################################### 로그인
                ##################################### 로그인
                ##################################### 로그인
                ##################################### 로그인

                # ID
                input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#mall_id")))
                input_field.click()
                time.sleep(1)
                input_field.send_keys(Keys.CONTROL + "a")
                input_field.send_keys(Keys.BACKSPACE)
                driver.find_element(By.CSS_SELECTOR, "#mall_id").send_keys(cafe24_id)

                # PW
                input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#userpasswd")))
                input_field.click()
                input_field.send_keys(Keys.CONTROL + "a")
                input_field.send_keys(Keys.BACKSPACE)
                driver.find_element(By.CSS_SELECTOR, "#userpasswd").send_keys(cafe24_pw)

                # 로그인클릭
                driver.find_element(By.CSS_SELECTOR,'#frm_user > div > div.mButton > button').click()

                #비밀번호변경안내
                try: WebDriverWait(driver, 5).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#iptBtnEm")))).click() 
                except: pass

                try:
                    time.sleep(3)
                    popup = driver.find_element(By.XPATH, '//*[contains(text(), "오늘 하루 보지 않기")]')
                    popup.click()

                except: pass

                #화면로딩대기
                WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "오늘의 할 일")]')))

                # 데이터 접근
                driver.get(url_cafe24_req)

                # 자세히보기클릭
                element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#QA_day3 > div.mBoard.gScroll > table")))
                driver.execute_script("arguments[0].scrollIntoView(true);", element) # 스크롤다운
                WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#sReportGabView"))).click() 
                
                element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#QA_day3 > div.mBoard.gScroll > table")))
                driver.execute_script("arguments[0].scrollIntoView(true);", element) # 스크롤다운
                rows = driver.find_elements(By.CSS_SELECTOR, 'tbody.right tr')

                cover = []
                cover0 = []
                for element in rows:
                    new_data_list = []
                    rawdata = element.text
                    # 문자열을 공백을 기준으로 분리하여 리스트로 변환
                    data_list = rawdata.split()
                    data_list = [x.replace(',', '') for x in data_list]

                    for items in data_list:

                    # 숫자인 경우 숫자로 변환
                        try:
                            numeric_value = int(items)
                            new_data_list.append(numeric_value)
                        except:
                            # 숫자가 아닌 경우 원래 값 유지
                            new_data_list.append(items)
                    cover.append(new_data_list[1:])
                    cover0.append(new_data_list[0])

                cover0_date_only = []

                # 입력할 데이터 정리
                for i in cover0:
                    cover0_date_only.append(i.split("(")[0])


                today_tdayTempDay = today_tday.strftime(f"%Y-%m-%d({weekday_krt})")
                cover.reverse()
                cover0_date_only.reverse()

                today_tdayTemp = today-timedelta(days=target_days_input)
                print(cover0_date_only)

                # 서비스 계정 키 파일 경로
                credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                # gspread 클라이언트 초기화
                client = gspread.service_account(filename=credential_file)

                # Google 시트 열기
                spreadsheet = client.open_by_url(sheet_urlR)

                # 첫 번째 시트 선택
                sheet = spreadsheet.worksheet(sheet_nameR)

                column_values = sheet.col_values(1)

                for i in range(target_days_input):
                    print(today_tdayTemp)

                    date_to_count = str(today_tdayTemp)
                    sheetCount = column_values.count(date_to_count)
                    listCount = cover0_date_only.count(date_to_count)

                    if str(today_tdayTemp) in column_values:
                        # 날짜 넘김 처리
                        today_tdayTemp = today_tdayTemp + timedelta(days=1)
                        continue

                    else:
                        last_row = len(sheet.get_all_values())
                        next_row = last_row + 1  # 다음 행 번호

                        print(next_row)
                        print(today_tdayTempDay)

                        # last_row = len(sheet.col_values(3))
                        # next_row = last_row + 1
                        # print(last_row)
                        # print(next_row)

                        if str(today_tdayTemp) in cover0_date_only:
                            print("성립")
                            keynum = cover0_date_only.index(str(today_tdayTemp))
                            data_to_paste = cover[keynum]  

                            data1 = data_to_paste[:9]
                            data2 = data_to_paste[9]
                            data3 = data_to_paste[10:]
                        # 카페24 R, 데이터 없으면 0 입력 되도록 코드 수정 -2
                        else:
                            data1 = [0, 0, 0, 0, 0, 0, 0, 0, 0]
                            data2 = 0
                            data3 = [0, 0, 0]
                            # data_date = 


                        print(data1)
                        print(data2)
                        print(data3)

                        if brand == "천명연구소":
                            range1 = f'C{next_row}:K{next_row}'
                            range2 = f'O{next_row}'
                            range3 = f'P{next_row}:Q{next_row}'

                        else:
                            range1 = f'C{next_row}:K{next_row}'
                            range2 = f'M{next_row}'
                            range3 = f'O{next_row}:Q{next_row}'
                        
                        sheet.update([data1], range1)
                        sheet.update([[data2]], range2)
                        sheet.update([data3], range3)
                        # 브랜드 이름 넣기
                        range_brand = f'B{next_row}'
                        sheet.update([[brand]], range_brand)
                        # 브랜드 이름 넣기
                        range_date = f'A{next_row}'
                        sheet.update([[str(today_tdayTemp)]], range_date)

                        # 날짜 넘김 처리
                        today_tdayTemp = today_tdayTemp + timedelta(days=1)
                        

                driver.close()

                # 완료 로그
                self.logBox.append(f"{self.sales_group_box.title()}-{self.salesCafe24.text()}-{brand}<br><span style='color:blue;'>완료</span>")

            except Exception as e:
                print(f"Error occurred: {e}")
                if driver:
                    try:
                        driver.close()  # 오류 발생 시 드라이버를 닫음
                        print("Driver closed successfully.")
                    except Exception as close_error:
                        print(f"Error closing driver: {close_error}")

                # 실패 로그
                self.logBox.append(f"{self.sales_group_box.title()}-{self.salesCafe24.text()}-{brand}<br><span style='color:red;'>실패</span>")

        sheet_sales_url = 'https://docs.google.com/spreadsheets/d/18dewLYnVKwTy9PTgYDqRurJBpVpXEVWZxrPrD3wJPg0/edit?gid=729300904#gid=729300904'
        sheet_advt_know_url = 'https://docs.google.com/spreadsheets/d/1CT15kvW9-ZLCJZNXrSsAe07eY9HLH2NTDRaDCqQU1h8/edit?gid=330152092#gid=330152092'
        sheet_advt_zenique_url = 'https://docs.google.com/spreadsheets/d/1U4s9UbjElH1QUk4-GvtTvxmHvzpswl22S8DWWkoWG9w/edit?gid=928641371#gid=928641371'
        sheet_advt_CMlabs_url = 'https://docs.google.com/spreadsheets/d/1qQf3ejDaRApLdOvPA6e8PONjIK6v-pFaXidPTB-1ZuY/edit?gid=973452544#gid=973452544'


        #카페24 천명연구소
        if self.CMlabs_salesCafe24.isChecked() == True:

            url_cafe24 = "https://eclogin.cafe24.com/Shop/" 
            url_cafe24_req_CMlabs = "https://cheonmyeong.cafe24.com/disp/admin/shop1/report/DailyList"

            cafe24_id_CMlabs = self.login_info("CAFE_CMlabs_ID")
            cafe24_pw_CMlabs = self.login_info("CAFE_CMlabs_PW")

            sheet_CMlabsR = '천명연구소R'
            # sheet_knowD = "노마셀D"
            brand = "천명연구소"

            cafe24(url_cafe24, url_cafe24_req_CMlabs, cafe24_id_CMlabs, cafe24_pw_CMlabs, sheet_sales_url, sheet_CMlabsR, brand)

        #카페24 노마셀
        if self.know_salesCafe24.isChecked() == True:

            url_cafe24 = "https://eclogin.cafe24.com/Shop/" 
            url_cafe24_req_knowmycell = "https://fkark12.cafe24.com/disp/admin/shop1/report/DailyList"

            cafe24_id_knowmycell = self.login_info("CAFE_KNOW_ID")
            cafe24_pw_knowmycell = self.login_info("CAFE_KNOW_PW")

            sheet_knowR = '노마셀R'
            sheet_knowD = "노마셀D"
            brand = "노마셀"

            cafe24(url_cafe24, url_cafe24_req_knowmycell, cafe24_id_knowmycell, cafe24_pw_knowmycell, sheet_sales_url, sheet_knowR, brand)

        #카페24 제니크
        if self.zq_salesCafe24.isChecked() == True:

            url_cafe24 = "https://eclogin.cafe24.com/Shop/" 
            url_cafe24_req_ZQ = "https://fkark08.cafe24.com/disp/admin/shop1/report/DailyList"

            cafe24_id_ZQ = self.login_info("CAFE_ZQ_ID")
            cafe24_pw_ZQ = self.login_info("CAFE_ZQ_PW")

            sheet_ZQR = '제니크R'
            sheet_ZQD = "제니크D"
            brand = "제니크"

            cafe24(url_cafe24, url_cafe24_req_ZQ, cafe24_id_ZQ, cafe24_pw_ZQ, sheet_sales_url, sheet_ZQR, brand)
    

# 쿠팡 매출
        def sales_coup(url, id, pw, sheet_url, sheet_name, option, brand):
            driver = None
            # try:

            # 크롬 On
            ### chromedriver_autoinstaller.install() 사용 추가
            chromedriver_path = chromedriver_autoinstaller.install()
            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_argument("--start-maximized") #최대 크기로 시작
            # chrome_options.add_argument('--incognito')
            # chrome_options.add_argument('--window-size=1920,1080')  
            # chrome_options.add_argument('--headless')
            chrome_options.add_experimental_option('detach', True)

            user_data = self.chrome_path_folder.text()
            chrome_options.add_argument(f"user-data-dir={user_data}")
            chrome_options.add_argument("--profile-directory=Profile 1")
            
            user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
            headers = {'user-agent' : user_agent}

            driver = webdriver.Chrome(
                service=Service(chromedriver_path),
                options=chrome_options
            )

            driver.get(url)

            print(id)
            print(pw)
            input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#username")))
            input_field.click()
            time.sleep(0.7)
            input_field.send_keys(Keys.CONTROL + "a")
            input_field.send_keys(Keys.BACKSPACE)
            driver.find_element(By.CSS_SELECTOR, "#username").send_keys(id)
            input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#password")))
            input_field.click()
            input_field.send_keys(Keys.CONTROL + "a")
            input_field.send_keys(Keys.BACKSPACE)
            driver.find_element(By.CSS_SELECTOR, "#password").send_keys(pw)
            driver.find_element(By.CSS_SELECTOR,'#kc-login').click()

            target_days = target_days_input
            

            while target_days > 0:

                # 서비스 계정 키 파일 경로
                credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                # gspread 클라이언트 초기화
                client = gspread.service_account(filename=credential_file)

                # Google 시트 열기
                spreadsheet = client.open_by_url(sheet_url)

                # 첫 번째 시트 선택
                sheet = spreadsheet.worksheet(sheet_name)

                last_row = len(sheet.get_all_values())
                print(last_row)
                next_row = last_row + 1  # 다음 행 번호

                # # 쿠팡 날짜 검색을 위해 unix 타임스탬프로 날짜 변경
                # def date_to_unix_timestamp(date_str):
                #     """
                #     주어진 날짜 문자열 (YYYY-MM-DD)을 Unix 타임스탬프 (밀리초)로 변환하는 함수.
                #     """
                #     dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")  # 문자열을 datetime 객체로 변환
                #     return int(dt.timestamp() * 1000)  # 초 단위를 밀리초(ms)로 변환


                dayx = datetime.timedelta(days=target_days)
                today_tday_str = (today-dayx).strftime('%Y-%m-%d')
                # today_tday_str_to_Unix = date_to_unix_timestamp(today_tday_str) // 쿠팡 사이트 변경으로 today_tday_str 사용

                if str(today_tday_str) in sheet.col_values(1):
                    target_days -= 1
                    continue

                # 다운로드 확인
                cnt = 1
                current_file_count1 = count_files(download_folder)
                while cnt < 10:

                
                    try:
                        try:
                            #날짜 선택
                            time.sleep(1)
                            # driver.get(f"https://wing.coupang.com/tenants/rfm-ss/business-insight/sales-analysis?startDate={today_tday_str}&endDate={today_tday_str}")
                            driver.get(f"https://wing.coupang.com/tenants/business-insight/sales-analysis?start_date={today_tday_str}&end_date={today_tday_str}")


                            #상품별 엑셀 다운로드
                            element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,f"//*[text()='엑셀 다운로드']")))
                            element.click()

                            element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,f"//*[text()='상품별 엑셀 다운로드']")))
                            element.click()

                        except: pass
                        
                    except:
                        try:

                            driver.get(url)

                            #날짜 선택
                            time.sleep(1)
                            # driver.get(f"https://wing.coupang.com/tenants/rfm-ss/business-insight/sales-analysis?startDate={today_tday_str}&endDate={today_tday_str}")
                            driver.get(f"https://wing.coupang.com/tenants/business-insight/sales-analysis?start_date={today_tday_str}&end_date={today_tday_str}")
                            

                            #상품별 엑셀 다운로드
                            element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,f"//*[text()='엑셀 다운로드']")))
                            element.click()

                            element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,f"//*[text()='상품별 엑셀 다운로드']")))
                            element.click()

                        except: pass

                    time.sleep(5)

                    current_file_count2 = count_files(download_folder)
                    if current_file_count1 != current_file_count2:
                        break

                    cnt += 1

                
                
                time.sleep(1)

                xlsx_file = get_latest_file(download_folder)

                df_uploaded_new = pd.read_excel(xlsx_file)
                # '러브슬라임'이라는 단어가 포함된 모든 행을 '상품명/Product Name/注册产品名称' 열을 기준으로 필터링합니다.
                filtered_rows_with_loveslime = df_uploaded_new[df_uploaded_new['상품명'].astype(str).str.contains(option)]

                # 필터링된 행들의 데이터를 리스트로 변환합니다.
                rows_list_with_loveslime = filtered_rows_with_loveslime.values.tolist()


                # 두 번째 값만 정수형으로 변환한 후 문자열로 변환하여 업데이트하는 과정
                updated_data_list = []
                for row in rows_list_with_loveslime:
                    new_row = row.copy()  # 원본 데이터의 복사본 생성
                    if len(row) > 1:  # 두 번째 값이 존재하는지 확인
                        new_row[0] = str(int(row[0]))  # 첫 번째 값을 정수형으로 변환 후 문자열로 변환
                    updated_data_list.append(new_row)

                # 결과 출력
                print(updated_data_list)

                
                    
                # Google 시트에 데이터 쓰기
                if len(updated_data_list) > 0:
                    i = 0
                    while i < len(updated_data_list):
                        range_to_write_Option_and_OptionId = f'C{next_row+i}:D{next_row+i}'
                        range_to_write_ProductId = f'B{next_row+i}'
                        range_to_write_ProductType_and_Category = f'E{next_row+i}:F{next_row+i}'
                        range_to_write_net_sales_amount = f'H{next_row+i}:I{next_row+i}'
                        range_to_write_total_sales_amount = f'J{next_row+i}:N{next_row+i}'

                        sheet.update([[today_tday_str]], f'A{next_row+i}')
                        print([updated_data_list[i][:2]])
                        sheet.update([updated_data_list[i][:2]], range_to_write_Option_and_OptionId)
                        sheet.update([[updated_data_list[i][3]]], range_to_write_ProductId)
                        temp_list = updated_data_list[i][4:6]
                        temp_list[0], temp_list[1] = temp_list[1], temp_list[0]
                        sheet.update([temp_list], range_to_write_ProductType_and_Category)
                        sheet.update([updated_data_list[i][6:8]], range_to_write_net_sales_amount)
                        sheet.update([updated_data_list[i][14:19]], range_to_write_total_sales_amount)
                        i += 1
                else:
                    dummy_data = ['-', '-', '-', '-', '-', '-', 0, 0, 0, 0, 0, 0, 0]
                    range_to_write = f'B{next_row}:N{next_row}'
                    sheet.update([dummy_data], range_to_write)
                    sheet.update([[today_tday_str]], f'A{next_row}')

                target_days -= 1
                
            driver.close()

            #     # 완료 로그
            #     self.logBox.append(f"{self.sales_group_box.title()}-{self.salesCoup.text()}-{brand}<br><span style='color:blue;'>완료</span>")

            # except Exception as e:
            #     print(f"Error occurred: {e}")
            #     if driver:
            #         try:
            #             driver.close()  # 오류 발생 시 드라이버를 닫음
            #             print("Driver closed successfully.")
            #         except Exception as close_error:
            #             print(f"Error closing driver: {close_error}")

            #     # 실패 로그
            #     self.logBox.append(f"{self.sales_group_box.title()}-{self.salesCoup.text()}-{brand}<br><span style='color:red;'>실패</span>")

        coupC_url = "https://wing.coupang.com/seller/notification/metrics/dashboard"

        # 쿠팡 노마셀
        if self.know_salesCoup.isChecked() == True:
            coupang_id_know = self.login_info("COUP_KNOW_ID")
            coupang_pw_know = self.login_info("COUP_KNOW_PW")
            sheet_name_knowC = '노마셀C'
            options = "노마셀"
            brand = "노마셀"


            sales_coup(coupC_url, coupang_id_know, coupang_pw_know, sheet_sales_url, sheet_name_knowC, options, brand)

# 네이버 매출
        def sales_naver(url, brand, sheet_name, sheet_url, brand_log):

            edge_driver = None
            try:

                # 날짜 구하기
                today = date.today()
                today_date = today.strftime("%d")
                today_Ym = today.strftime("%Y. %m.")

                number = target_days_input
                dayx = datetime.timedelta(days=number)
                dayy = datetime.timedelta(days=1)
                day1 = datetime.timedelta(days=1)

                today_yday = today - day1
                startday = today - dayx
                endday = today - dayy
                tday_Ym = startday.strftime("%Y. %m.")
                tday_d = startday.strftime("%d")

                # EdgeOptions 설정
                edge_options = webdriver.EdgeOptions()
                edge_options.use_chromium = True
                edge_options.add_argument("disable-gpu")
                edge_options.add_argument("no-sandbox")
                edge_options.add_argument(f"user-data-dir={self.edge_path_folder.text()}")
                edge_options.add_argument("--profile-directory=Default")

                # 현재 실행 파일 기준 디렉토리
                if getattr(sys, 'frozen', False):
                    # PyInstaller로 패키징된 실행파일이면 실행파일 위치 기준
                    base_path = sys._MEIPASS if hasattr(sys, '_MEIPASS') else os.path.dirname(sys.executable)
                else:
                    # 개발환경(py로 실행)에서는 현재 경로
                    base_path = os.path.dirname(os.path.abspath(__file__))

                # 드라이버 경로 지정 (같은 폴더에 넣었다고 가정)
                driver_path = os.path.join(base_path, "msedgedriver.exe")

                if not os.path.exists(driver_path):
                    raise FileNotFoundError(f"Edge 드라이버가 없습니다: {driver_path}")

                # 드라이버 실행
                edge_service = Service(driver_path)
                edge_driver = webdriver.Edge(service=edge_service, options=edge_options)

                edge_driver.get(url)

                # 로그인
                WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#wrap > div > div > div.login_box > ul > li:nth-child(1) > a"))).click()
                try:
                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#root > div > div.Layout_wrap__9yckO > div > div > div.Login_simple_box__zyz-B > button"))).click()
                
                except:
                    edge_driver.find_element(By.CSS_SELECTOR, '[class^="Login_btn_more"]').click()

                    current_window_handle = edge_driver.current_window_handle

                    new_window_handle = None
                    while not new_window_handle:
                        for handle in edge_driver.window_handles:
                            if handle != current_window_handle:
                                new_window_handle = handle
                                break

                    #팝업으로 제어 변경
                    edge_driver.switch_to.window(edge_driver.window_handles[1])


                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#log\.login")))
                    
                    txtInput = edge_driver.find_element(By.CSS_SELECTOR, "#id")
                    txtInput.send_keys("wntlsqhr")
                    time.sleep(0.1)
                    txtInput = edge_driver.find_element(By.CSS_SELECTOR, "#pw")
                    txtInput.send_keys("dnflskfk00@")
                    time.sleep(0.1)
                    edge_driver.find_element(By.CSS_SELECTOR, "#log\.login")

                    #원래 페이지로 제어 변경
                    edge_driver.switch_to.window(edge_driver.window_handles[0])


                # 상품별 이동
                WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#include_nav > div > div > div:nth-child(1) > ul > li:nth-child(4) > a"))).click()
                WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#include_nav > div > div > div:nth-child(1) > ul > li.on > div > ul > li:nth-child(1) > a"))).click()

                brandtext = edge_driver.find_element(By.CSS_SELECTOR, "#include_header > div > div.header_tit > div > div:nth-child(2) > div > div > div > span:nth-child(1)").text[:3]


                # 브랜드 변경
                if not brandtext == brand:
                    edge_driver.find_element(By.CSS_SELECTOR, "#include_header > div > div.header_tit > div > div:nth-child(2) > div > div > div > div > a > div > span").click()

                while startday != today:

                    # 날짜 클릭(달력오픈)
                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.date_select > a.btn.select_data'))).click()

                    # 날짜 변수 지정
                    tday_Ym = startday.strftime("%Y. %m.")
                    tday_d = str(int(startday.strftime("%d")))
                    trick = (startday-day1).strftime("%Y. %m.")
                    print(startday)

                    # %Y. %m 표시(웹상)
                    DPmonthStart = edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker.DayPicker > div > div.DayPicker-Month.rdp-caption_start")
                    DPmonthBtw = edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker.DayPicker > div > div.DayPicker-Month.rdp-caption_between")
                    DPmonthEnd = edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on.open > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker.DayPicker > div > div.DayPicker-Month.rdp-caption_end")


                    # 첫번째 단락 년,월 대조
                    if tday_Ym == DPmonthStart.text[:9]:

                            days = edge_driver.find_elements(By.XPATH, f"//*[@id='wrap']/div[1]/section/div/div[2]/div[1]/div/ul/li[2]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/table/tbody//td[not(contains(@class, 'DayPicker-Day--outside'))]")
                            for day in days:
                                if day.text == tday_d:
                                    print("target: ", day.text)
                                    day.click()
                                    time.sleep(0.1)
                                    day.click()
                                    break

                            # 적용
                            edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_info_area > div.btn_area > a:nth-child(1)").click()
                    
                    
                    # 두번째 단락 년,월 대조
                    elif tday_Ym == DPmonthBtw.text[:9]:
                    
                            days = edge_driver.find_elements(By.XPATH, f"//*[@id='wrap']/div[1]/section/div/div[2]/div[1]/div/ul/li[2]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[2]/table/tbody//td[not(contains(@class, 'DayPicker-Day--outside'))]")
                            for day in days:
                                if day.text == tday_d:
                                    print("target: ", day.text)
                                    day.click()
                                    time.sleep(0.1)
                                    day.click()
                                    break
                                
                            # 적용
                            edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_info_area > div.btn_area > a:nth-child(1)").click()

                    # 세번째 단락 년,월 대조 /// 250602추가
                    elif tday_Ym == DPmonthEnd.text[:9]:
                    
                            days = edge_driver.find_elements(By.XPATH, f"//*[@id='wrap']/div[1]/section/div/div[2]/div[1]/div/ul/li[2]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[3]/table/tbody//td[not(contains(@class, 'DayPicker-Day--outside'))]")
                            for day in days:
                                if day.text == tday_d:
                                    print("target: ", day.text)
                                    day.click()
                                    time.sleep(0.1)
                                    day.click()
                                    break
                                
                            # 적용
                            edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_info_area > div.btn_area > a:nth-child(1)").click()

                    # 5. 네이버 스스 이전 달 날짜 선택 안되는 코드 변경
                    else:
                        # 이전 달로 이동
                        WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker-NavBar > span.DayPicker-NavButton.DayPicker-NavButton--prev'))).click()

                        if tday_Ym == DPmonthStart.text[:9]:

                            days = edge_driver.find_elements(By.XPATH, f"//*[@id='wrap']/div[1]/section/div/div[2]/div[1]/div/ul/li[2]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/table/tbody//td[not(contains(@class, 'DayPicker-Day--outside'))]")
                            for day in days:
                                if day.text == tday_d:
                                    print("target: ", day.text)
                                    day.click()
                                    time.sleep(0.1)
                                    time.sleep(2)
                                    day.click()
                                    break

                            # 적용
                            time.sleep(2)
                            edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_info_area > div.btn_area > a:nth-child(1)").click()
                    
                
                    # 다운로드 버튼
                    element = WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(1) > span > a')))

                    # 다운로드 확인
                    cnt = 1
                    current_file_count1 = count_files(download_folder)
                    while cnt < 10:
                        try:
                            element.click()
                        except:
                            print("error")

                        time.sleep(3)
                        current_file_count2 = count_files(download_folder)
                        if current_file_count1 != current_file_count2:
                            break

                        cnt += 1

                    # check_download()
                    time.sleep(1)

                    startday += datetime.timedelta(days=1)
                    

                time.sleep(2)

                try:
                    edge_driver.close()
                except (InvalidSessionIdException, WebDriverException) as e:
                    print("브라우저 세션이 이미 종료되었거나 무효함:", e)

                defaultData = ["-", "-", "-", "-", "-", "-", "-", "0", "0", "0", "0.00%"]
                # 날짜 구하기
                today = date.today()

                today_date = today.strftime("%d")
                today_Ym = today.strftime("%Y. %m.")

                number = target_days_input
                dayx = datetime.timedelta(days=number)
                dayy = datetime.timedelta(days=1)
                day1 = datetime.timedelta(days=1)

                today_yday = today-day1
                startday = today-dayx
                endday = today-dayy
                tday_Ym = startday.strftime("%Y. %m.")
                tday_d = startday.strftime("%d")

                while number > 0:

                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'
                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)
                    # Google 시트 열기
                    spreadsheet = client.open_by_url(sheet_url)
                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(sheet_name)

                    last_row = len(sheet.col_values(1))
                    print(last_row)
                    next_row = last_row + 1  # 다음 행 번호

                    if str(startday) in sheet.col_values(1):
                        number -= 1
                        startday += timedelta(days=1)
                        continue

                    i = get_nth_latest_file(download_folder, number)

                    wb = openpyxl.load_workbook(i)
                    sheet = wb.active  # 활성 시트 선택

                    if check_data_in_second_row(i):
                        pass

                    else:
                        sheet = spreadsheet.worksheet(sheet_name)

                        # 날짜 입력
                        sheet.update([[str(startday)]], f"A{next_row}")
                        range_to_write = f'B{next_row}:L{next_row}'
                        sheet.update([defaultData], range_to_write)
                        number -= 1
                        startday += timedelta(days=1)  # 날짜 하루 증가
                        continue


                    # 원본 시트의 행을 반복하며 첫 번째 행을 제외하고 데이터가 있는 행만 복사
                    for row in sheet.iter_rows(min_row=2):  # 첫 번째 행을 제외하고 시작
                        # 각 셀에 데이터가 있는지 확인
                        data_exists = any(cell.value not in (None, '', ' ') for cell in row)  # 빈 문자열과 공백도 무시

                        # 서비스 계정 키 파일 경로
                        credential_file = 'triple-nectar-412808-da4dac0cc16e.json'
                        # gspread 클라이언트 초기화
                        client = gspread.service_account(filename=credential_file)
                        # Google 시트 열기
                        spreadsheet = client.open_by_url(sheet_url)
                        # 첫 번째 시트 선택
                        sheet = spreadsheet.worksheet(sheet_name)

                        # 날짜 입력
                        sheet.update([[str(startday)]], f"A{next_row}")
                        
                        values = []
                        for col_index, cell in enumerate(row, start=2):
                            values.append(cell.value)

                        range_to_write = f'B{next_row}:L{next_row}'
                        sheet.update([values], range_to_write)

                        
                        next_row += 1

                    startday += timedelta(days=1)  # 날짜 하루 증가
                    number -= 1

                # 완료 로그
                self.logBox.append(f"{self.sales_group_box.title()}-{self.salesNaver.text()}-{brand_log}<br><span style='color:blue;'>완료</span>")

                

            except Exception as e:
                print(f"Error occurred: {e}")
                if edge_driver:
                    try:
                        edge_driver.close()  # 오류 발생 시 드라이버를 닫음
                        print("Driver closed successfully.")
                    except Exception as close_error:
                        print(f"Error closing driver: {close_error}")
                self.logBox.append(f"{self.sales_group_box.title()}-{self.salesNaver.text()}-{brand}<br><span style='color:red;'>실패</span>")

        # def ssWrite(sheet_name, sheet_url, brand_log):

        url = "https://bizadvisor.naver.com/shopping/product"

        if self.know_salesNaver.isChecked() == True:

            brand = "노마셀"
            sheet_name = "노마셀N"
            brand_log = "노마셀"

            sales_naver(url, brand, sheet_name, sheet_sales_url, brand_log)


# 쿠팡 광고
        def advt_coupang(prepUrl, url_coupang_daily, id, pw, sheet_url, sheet_name, brand):
            dummyData = ["-", "-", "-", 0, "-", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "-"]

            driver = None
            try:
                # 크롬 On
                ### chromedriver_autoinstaller.install() 사용 추가
                chromedriver_path = chromedriver_autoinstaller.install()
                chrome_options = webdriver.ChromeOptions()
                chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
                chrome_options.add_argument("--start-maximized") #최대 크기로 시작
                # chrome_options.add_argument('--incognito')
                # chrome_options.add_argument('--window-size=1920,1080')  
                # chrome_options.add_argument('--headless')
                chrome_options.add_experimental_option('detach', True)

                user_data = self.chrome_path_folder.text()
                chrome_options.add_argument(f"user-data-dir={user_data}")
                chrome_options.add_argument("--profile-directory=Profile 1")
                
                user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
                headers = {'user-agent' : user_agent}

                driver = webdriver.Chrome(
                    service=Service(chromedriver_path),
                    options=chrome_options
                )


                try:
                    driver.get(prepUrl)  # 로그인 시작
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="main-container"]/div/div[1]/ul/li[1]/a/span'))).click()

                except NoSuchElementException:
                    # 요소가 없을 때 처리할 로직
                    pass


                print(id)
                print(pw)
                input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#username")))
                input_field.click()
                time.sleep(0.7)
                input_field.send_keys(Keys.CONTROL + "a")
                input_field.send_keys(Keys.BACKSPACE)
                driver.find_element(By.CSS_SELECTOR, "#username").send_keys(id)
                input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#password")))
                input_field.click()
                input_field.send_keys(Keys.CONTROL + "a")
                input_field.send_keys(Keys.BACKSPACE)
                driver.find_element(By.CSS_SELECTOR, "#password").send_keys(pw)
                driver.find_element(By.CSS_SELECTOR,'#kc-login').click()

                # 로그인 오류 발생하면 재시도
                ### 비밀번호 오류 예외문
                try:
                    loginErrorMessage = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "아이디 또는 비밀번호가 다릅니다.")]')))
                    if loginErrorMessage:
                        input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#username")))
                        input_field.click()
                        time.sleep(0.7)
                        input_field.send_keys(Keys.CONTROL + "a")
                        input_field.send_keys(Keys.BACKSPACE)
                        driver.find_element(By.CSS_SELECTOR, "#username").send_keys(id)
                        input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#password")))
                        input_field.click()
                        input_field.send_keys(Keys.CONTROL + "a")
                        input_field.send_keys(Keys.BACKSPACE)
                        driver.find_element(By.CSS_SELECTOR, "#password").send_keys(pw)
                        driver.find_element(By.CSS_SELECTOR,'#kc-login').click()
                    

                except: pass

                try:
                    driver.get(prepUrl)  # 로그인 시작
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[contains(text(), "광고 관리")]')))
                    print("3")
                    driver.get(url_coupang_daily)  # 로그인 시작
                    if driver.find_element(By.CSS_SELECTOR, "body > h1"):
                        driver.get(prepUrl)  # 로그인 시작
                        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[contains(text(), "광고 관리")]')))
                        print("4")
                        driver.get(url_coupang_daily)  # 요소가 존재하면 페이지를 다시 로드
                except NoSuchElementException:
                    pass

                try:
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "기간 설정")]'))).click()
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CLASS_NAME,"ant-picker-input-active"))).click() #클릭 시작일
                except:
                    print("보고서페이지 로딩실패... retry")
                    driver.find_element(By.CSS_SELECTOR, "#cap-sidebar > nav > ul > li.ant-menu-item.ant-menu-item-selected > span.ant-menu-title-content > span").click()

                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#ad-reporting-app > div.self-service-ad-reporting-ui > div > div.sc-ert4fh-2.eLreXy > div > div.tabs > div:nth-child(2) > div"))).click()

                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "매출 성장 광고 보고서")]'))).click()


                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "기간 설정")]'))).click()
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CLASS_NAME,"ant-picker-input-active"))).click() #클릭 시작일

                before_Ym = today_tday.strftime("%Y년 %m월")
                before_d = str(int(today_tday.strftime("%d")))
                yesterday_Ym = today_yday.strftime("%Y년 %m월")
                yesterday_d = str(int(today_yday.strftime("%d")))

                # 시작 날짜 입력
                input_field1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[placeholder='시작일']")))
                print("1")

                #ad-reporting-app > div.self-service-ad-reporting-ui > div > div.sc-ert4fh-2.eLreXy > div > div.sc-11l2gxs-0.brOYQE > div.sc-ipia07-0.iCqAxH > div.panel-options > div:nth-child(4) > div > div > div > div > div:nth-child(4) > div > div.ant-picker-input.ant-picker-input-active > input
                input_field1.send_keys(str(today_tday))
                time.sleep(0.5)
                input_field1.send_keys(Keys.ENTER)
                print("2")
                
                # 종료 날짜 입력
                actions = ActionChains(driver)
                print("3")
                actions.send_keys(f"{str(today_yday)}").perform()
                print("4")
                time.sleep(0.7)
                print("5")
                actions.send_keys(Keys.ENTER).perform()
                print("6")

                time.sleep(1)
                # element = driver.find_element(By.CSS_SELECTOR, "input[value='daily']")#기간 구분 // 아랫줄이랑 같은기능(time wait 이 빠진)
                element = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[value='daily']")))#기간 구분
                print("7")
                element.click() 
                print("8")
                # ActionChains(driver).move_to_element_with_offset(element,5,75).click().perform() #클릭 일별
                time.sleep(0.3)
                
                WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CLASS_NAME, 'campaign-picker-dropdown-btn'))).click() # 캠페인 선택
                time.sleep(0.5)
                # 전체선택 체크박스
                WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "전체선택")]'))).click()
                # 확인버튼 클릭
                WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.ant-btn.ant-btn-primary.confirm-button'))).click()
                time.sleep(0.3)

                ### 보고서 생성 실패하면 페이지 다시 로딩 후 생성
                xpath = '//*[@id="rc-tabs-0-panel-requestedReport"]/div[1]/div/div/div[3]/div[2]/div[2]/div[4]/div[1]/div[2]/div/div[10]/div[6]/div/button[2]/span'
                try:
                    try:
                        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), '보고서 만들기')]"))).click()

                    except:
                        driver.find_element(By.CLASS_NAME,'campaign-picker-dropdown-btn').click() #캠페인 선택
                        time.sleep(0.5)
                        # 전체선택 체크박스
                        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "전체선택")]'))).click()
                        # 확인버튼 클릭
                        driver.find_element(By.CSS_SELECTOR, '.ant-btn.ant-btn-primary.confirm-button').click()
                        time.sleep(0.3)
                        # 보고서생성
                        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), '보고서 만들기')]"))).click()

                    time.sleep(5)

                    # if driver.find_element(By.CSS_SELECTOR, "#rc-tabs-0-panel-requestedReport > div > div.react-grid-Container > div > div > div:nth-child(2) > div > div > div:nth-child(2) > div:nth-child(1) > div:nth-child(5) > div > div > span > div").text == "생성 실패":
                    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), '보고서 만들기')]"))).click()
                    #     time.sleep(5)

                    
                    # 보고서 다운로드
                    element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, f"{xpath}")))
                    
                    

                    # 다운로드 확인
                    cnt = 1
                    current_file_count1 = count_files(download_folder)
                    while cnt < 10:
                        try:
                            element.click()
                        except:
                            driver.find_element(By.CSS_SELECTOR, "body > div:nth-child(12) > div > div.ant-modal-wrap > div > div.ant-modal-content > div > div > div.ant-modal-confirm-btns > button").click()
                            element.click()
                        time.sleep(3)
                        current_file_count2 = count_files(download_folder)
                        if current_file_count1 != current_file_count2:
                            break

                        cnt += 1

                    # check_download()
                    time.sleep(1)
                    try:
                        WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "body > div.MuiDialog-root.sc-852clq-0.efPzRF > div.MuiDialog-container.MuiDialog-scrollPaper > div > div:nth-child(3) > button"))).click()
                    except: pass

                except:
                    driver.get(prepUrl)  # 로그인 시작
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[contains(text(), "광고 관리")]')))
                    print("5")
                    driver.get(url_coupang_daily)

                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "기간 설정")]'))).click()
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CLASS_NAME,"ant-picker-input-active"))).click() #클릭 시작일

                    # 시작 날짜 입력
                    input_field1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[placeholder='시작일']")))

                    input_field1.send_keys(str(today_tday))
                    time.sleep(0.5)
                    input_field1.send_keys(Keys.ENTER)
                    
                    # 종료 날짜 입력
                    actions = ActionChains(driver)
                    actions.send_keys(f"{str(today_yday)}").perform()
                    time.sleep(0.3)
                    actions.send_keys(Keys.ENTER).perform()

                    element = driver.find_element(By.CSS_SELECTOR, "input[value='daily']")#기간 구분
                    element.click() 
                    # ActionChains(driver).move_to_element_with_offset(element,5,75).click().perform() #클릭 일별
                    time.sleep(0.3)

                    driver.find_element(By.CLASS_NAME,'campaign-picker-dropdown-btn').click() #캠페인 선택
                    time.sleep(0.5)
                    # 전체선택 체크박스
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "전체선택")]'))).click()
                    # 확인버튼 클릭
                    driver.find_element(By.CSS_SELECTOR, '.ant-btn.ant-btn-primary.confirm-button').click()
                    time.sleep(0.3)

                    ### 보고서 생성 실패하면 페이지 다시 로딩 후 생성
                    try:
                        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), '보고서 만들기')]"))).click()

                    except:
                        driver.find_element(By.CLASS_NAME,'campaign-picker-dropdown-btn').click() #캠페인 선택
                        time.sleep(0.5)
                        # 전체선택 체크박스
                        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "전체선택")]'))).click()
                        # 확인버튼 클릭
                        driver.find_element(By.CSS_SELECTOR, '.ant-btn.ant-btn-primary.confirm-button').click()
                        time.sleep(0.3)
                        # 보고서생성
                        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), '보고서 만들기')]"))).click()

                    time.sleep(5)

                    # if driver.find_element(By.CSS_SELECTOR, "#rc-tabs-0-panel-requestedReport > div > div.react-grid-Container > div > div > div:nth-child(2) > div > div > div:nth-child(2) > div:nth-child(1) > div:nth-child(5) > div > div > span > div").text == "생성 실패":
                    #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), '보고서 만들기')]"))).click()
                    #     time.sleep(5)

                    

                    # 보고서 다운로드
                    element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, f"{xpath}")))


                    # 다운로드 확인
                    cnt = 1
                    current_file_count1 = count_files(download_folder)
                    while cnt < 10:
                        try:
                            element.click()
                        except:
                            driver.find_element(By.CSS_SELECTOR, "body > div:nth-child(12) > div > div.ant-modal-wrap > div > div.ant-modal-content > div > div > div.ant-modal-confirm-btns > button").click()
                            element.click()
                        time.sleep(3)
                        current_file_count2 = count_files(download_folder)
                        if current_file_count1 != current_file_count2:
                            break

                        cnt += 1

                    # check_download()
                    time.sleep(1)
                    try:
                        WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "body > div.MuiDialog-root.sc-852clq-0.efPzRF > div.MuiDialog-container.MuiDialog-scrollPaper > div > div:nth-child(3) > button"))).click()
                    except: pass

                driver.close()

                time.sleep(2)

                xlsx_file = get_latest_file(download_folder)

                df_uploaded_new = pd.read_excel(xlsx_file)
                # '러브슬라임'이라는 단어가 포함된 모든 행을 '옵션명' 열을 기준으로 필터링합니다.
                filtered_rows_with_loveslime = df_uploaded_new[df_uploaded_new['과금방식'].astype(str).str.contains("cpc")]


                # 필터링된 행들의 데이터를 리스트로 변환합니다.
                rows_list_with_loveslime = filtered_rows_with_loveslime.values.tolist()

                excel_dates = []
                for i in rows_list_with_loveslime:
                    excel_dates.append(i[0])
                print(excel_dates)

                formatted_excel_dates = [datetime.datetime.strptime(str(date), "%Y%m%d").strftime("%Y-%m-%d") for date in excel_dates]

                print(formatted_excel_dates)

                updated_data_list = []
                for row in rows_list_with_loveslime:
                    new_row = row.copy()  # 원본 데이터의 복사본 생성
                    if len(row) > 1:  # 두 번째 값이 존재하는지 확인
                        new_row[1] = str(row[1])  # 두 번째 값을 정수형으로 변환 후 문자열로 변환

                    else:
                        pass
                        # dummyData = ["-", "-", "-", 0, "-", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "-"]
                                
                    updated_data_list.append(new_row)

                # 소수점 6번째 자리에서 반올림
                for i in range(len(updated_data_list)):  # 행 인덱스
                    for j in range(len(updated_data_list[i])):  # 열 인덱스
                        if isinstance(updated_data_list[i][j], float):  # 값이 float이면 반올림
                            updated_data_list[i][j] = round(updated_data_list[i][j], 6)
                            
                for sub_list in updated_data_list:
                    del sub_list[2]

                # 서비스 계정 키 파일 경로
                credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                # gspread 클라이언트 초기화
                client = gspread.service_account(filename=credential_file)

                # Google 시트 열기
                spreadsheet = client.open_by_url(sheet_url)

                # 첫 번째 시트 선택
                sheet = spreadsheet.worksheet(sheet_name)

                print(updated_data_list)
                
                last_row = len(sheet.get_all_values())
                print(last_row)
                next_row = last_row + 1  # 다음 행 번호

                formatted_date = today_yday.strftime("%Y-%m-%d")
                # Google 시트에 데이터 쓰기

                def convert_list_types(input_list):
                    converted_list = []
                    for item in input_list:
                        # 숫자로 변환할 수 있는 경우 변환 (int, float)
                        if item.replace('.', '', 1).isdigit():  
                            # 소수점이 있는 경우 float, 없는 경우 int로 변환
                            converted_list.append(float(item) if '.' in item else int(item))
                        else:
                            # 변환할 수 없는 경우 그대로 유지 (문자열)
                            converted_list.append(item)
                    return converted_list

                if len(updated_data_list) > 0:
                    i = 0
                    cnt = 0

                    # 첫 번째 열의 모든 값 가져오기
                    column_values = sheet.col_values(1)

                    # 오늘 날짜와 일치하는 값 개수 세기
                    count_theDay = []
                    for ii, value in enumerate(column_values):
                        if value == formatted_date:
                            count_theDay.append(ii+1)

                    print(count_theDay)
                    print(len(count_theDay))

                    for ii in count_theDay:
                        listTemp = convert_list_types(sheet.get_values(f'F{ii}:AI{ii}')[0])

                        print(listTemp)
                        result = [row[5:-1] for row in updated_data_list]
                        print(result)

                        if listTemp in result:
                            index_to_remove = result.index(listTemp)  # listTemp의 위치 찾기
                            del updated_data_list[index_to_remove]  # updated_data_list에서 해당 위치 삭제
                    print("★★★", updated_data_list)

                        


                    while cnt < len(updated_data_list):
                        print((updated_data_list[cnt])[1:-1])
                        print(i)
                        print(next_row+i)
                        print(sheet.get_values(f'B{next_row+i}'))
                        print(sheet.get_values(f'B{next_row+i}:AI{next_row+i}'))

                        # 중복 입력 방지
                        if formatted_excel_dates[cnt] in sheet.col_values(1) and sheet.get_values(f'B{last_row+i}:AI{last_row+i}') == [(updated_data_list[cnt])[1:-1]]:
                            cnt += 1
                            i += 1
                            continue
                        range_to_write = f'B{next_row+i}:AI{next_row+i}'
                        sheet.update([(updated_data_list[cnt])[1:-1]], range_to_write)
                        sheet.update([[formatted_excel_dates[cnt]]], f'A{next_row+i}')
                        i += 1
                        cnt += 1

                else:
                    today_tdayTemp = today-timedelta(days=target_days_input)

                    for i in range(target_days_input):
                        if str(today_tdayTemp) in sheet.col_values(1):
                            print(today_tdayTemp, "pass")
                            today_tdayTemp += timedelta(days=1)
                            continue
                        else:
                            print(today_tdayTemp, "non - pass")
                            range_to_write = f'B{next_row}:AI{next_row}'
                            sheet.update([(dummyData)], range_to_write)
                            sheet.update([[str(today_tdayTemp)]], f'A{next_row}')
                        next_row += 1
                        today_tdayTemp += timedelta(days=1)

                # 완료 로그
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtCoup.text()}-{brand}<br><span style='color:blue;'>완료</span>")
                        

            except Exception as e:
                print(f"Error occurred: {e}")
                if driver:
                    try:
                        driver.close()  # 오류 발생 시 드라이버를 닫음
                        print("Driver closed successfully.")
                    except Exception as close_error:
                        print(f"Error closing driver: {close_error}")
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtCoup.text()}-{brand}<br><span style='color:red;'>실패</span>")

#########쿠팡로데이터##########
        # def advt_coupang_rawdata(sheet_url, sheet_name, brand):

        coupC_url = "https://wing.coupang.com/seller/notification/metrics/dashboard"
        prep_coup_report_url = 'https://advertising.coupang.com/marketing/dashboard/sales'
        coup_report_url = 'https://advertising.coupang.com/marketing-reporting/billboard/reports/pa'
        sheet_url_coupC = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=374561563'

        # 쿠팡 노마셀
        if self.know_advtCoup.isChecked() == True:
            coupang_id_knowmycell = self.login_info("COUP_KNOW_ID")
            coupang_pw_knowmycell = self.login_info("COUP_KNOW_PW")
            sheet_url_know_all = 'https://docs.google.com/spreadsheets/d/1CT15kvW9-ZLCJZNXrSsAe07eY9HLH2NTDRaDCqQU1h8/edit?gid=330152092#gid=330152092'
            sheet_name_knowR = '노마셀 쿠팡 R'
            brand = "노마셀"

            advt_coupang(prep_coup_report_url, coup_report_url, coupang_id_knowmycell, coupang_pw_knowmycell, sheet_advt_know_url, sheet_name_knowR, brand)


### 네이버 검색광고 광고
         
        def advt_naver(url1, url2, name, brand):

            edge_driver = None
            try:

                # EdgeOptions 설정
                edge_options = webdriver.EdgeOptions()
                edge_options.use_chromium = True
                edge_options.add_argument("disable-gpu")
                edge_options.add_argument("no-sandbox")
                edge_options.add_argument(f"user-data-dir={self.edge_path_folder.text()}")
                edge_options.add_argument("--profile-directory=Default")

                # 현재 실행 파일 기준 디렉토리
                if getattr(sys, 'frozen', False):
                    # PyInstaller로 패키징된 실행파일이면 실행파일 위치 기준
                    base_path = sys._MEIPASS if hasattr(sys, '_MEIPASS') else os.path.dirname(sys.executable)
                else:
                    # 개발환경(py로 실행)에서는 현재 경로
                    base_path = os.path.dirname(os.path.abspath(__file__))

                # 드라이버 경로 지정 (같은 폴더에 넣었다고 가정)
                driver_path = os.path.join(base_path, "msedgedriver.exe")

                if not os.path.exists(driver_path):
                    raise FileNotFoundError(f"Edge 드라이버가 없습니다: {driver_path}")

                # 드라이버 실행
                edge_service = Service(driver_path)
                edge_driver = webdriver.Edge(service=edge_service, options=edge_options)

                edge_driver.get(url1)


                try:
                    # 로그인
                    WebDriverWait(edge_driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#wrap > div > div > div.login_box > ul > li:nth-child(1) > a"))).click()
                    try:
                        WebDriverWait(edge_driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#root > div > div.Layout_wrap__9yckO > div > div > div.Login_simple_box__zyz-B > button"))).click()
                    
                    except:
                        edge_driver.find_element(By.CSS_SELECTOR, '[class^="Login_btn_more"]').click()

                        current_window_handle = edge_driver.current_window_handle

                        new_window_handle = None
                        while not new_window_handle:
                            for handle in edge_driver.window_handles:
                                if handle != current_window_handle:
                                    new_window_handle = handle
                                    break

                        #팝업으로 제어 변경
                        edge_driver.switch_to.window(edge_driver.window_handles[1])


                        WebDriverWait(edge_driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#log\.login")))
                        
                        txtInput = edge_driver.find_element(By.CSS_SELECTOR, "#id")
                        txtInput.send_keys("wntlsqhr")
                        time.sleep(0.1)
                        txtInput = edge_driver.find_element(By.CSS_SELECTOR, "#pw")
                        txtInput.send_keys("dnflskfk00@")
                        time.sleep(0.1)
                        edge_driver.find_element(By.CSS_SELECTOR, "#log\.login")

                        #원래 페이지로 제어 변경
                        edge_driver.switch_to.window(edge_driver.window_handles[0])
                except: pass
                

                # 네이버검색광고 로그인 확인 창 제거 로직 수정(visibility_of_all_elements_located -> element_to_be_clickable)
                try:
                    WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "환영합니다")]')))
                    edge_driver.find_element(By.CLASS_NAME, "btn_name").click()
                    print("로그인확인 창 제거")
                except: 
                    print("로그인확인 창 없음")
                    pass
                
                # 네이버검색광고 로그인 확인 창 제거 로직 수정(visibility_of_all_elements_located -> element_to_be_clickable)
                try:
                    WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "환영합니다")]')))
                    checkbox = edge_driver.find_element(By.ID, "chk_cls")
                    checkbox.click()
                

                except: pass


                # 캘린더 열기
                WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-nclick="datePicker"]'))).click()

                # 날짜 선택
                WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "지난 7일")]'))).click()

                time.sleep(0.5)
                element = edge_driver.find_element(By.XPATH, "//*[contains(text(), '다운로드') and not(contains(text(), '대용량 다운로드 보고서'))]")
                time.sleep(0.5)

                # 다운로드 확인
                cnt = 1
                current_file_count1 = count_files(download_folder)
                while cnt < 10:
                    element.click()
                    time.sleep(3)
                    current_file_count2 = count_files(download_folder)
                    if current_file_count1 != current_file_count2:
                        break

                    cnt += 1
                    
                
                time.sleep(2)

                try:
                    edge_driver.close()
                except (InvalidSessionIdException, WebDriverException) as e:
                    print("브라우저 세션이 이미 종료되었거나 무효함:", e)

                target_days = target_days_input
                dayx = datetime.timedelta(days=target_days)
                day1 = datetime.timedelta(days=1)

                # 오늘 날짜 구하기
                today_yday = today-day1
                today_tday = today-dayx
                # CSV 파일 읽기 (첫 번째 행은 건너뛰고 두 번째 행을 열 이름으로 사용)
                df = pd.read_csv(get_latest_file(download_folder), skiprows=1)

                # 날짜 열 이름 추출 (A열, 즉 첫 번째 열)
                date_column = df.columns[0]

                # 'date' 열을 datetime 형식으로 변환
                df[date_column] = pd.to_datetime(df[date_column], format='%Y.%m.%d.')

                df[date_column] = df[date_column].dt.strftime('%Y-%m-%d')
                # 필터링된 데이터프레임 출력
                dataList = df.values.tolist()
                print(df.values.tolist())

                csv_file = get_latest_file(download_folder)

                # 서비스 계정 키 파일 경로
                credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                # gspread 클라이언트 초기화
                client = gspread.service_account(filename=credential_file)

                # Google 시트 열기
                spreadsheet = client.open_by_url(url2)

                # 첫 번째 시트 선택
                sheet = spreadsheet.worksheet(name)

                while today_tday != today:
                    # 중복 입력 방지
                    print(today_tday, "검색 시작")
                    if str(today_tday) in sheet.col_values(1):
                        today_tday += timedelta(days=1)
                        continue

                    for i in dataList:
                        if str(today_tday) in i:
                            result = []
                            print(today_tday, "찾음!")

                            for item in i:
                                if isinstance(item, str) and '%' in item:
                                    result.append(float(item.strip('%')) / 100)
                                elif isinstance(item, str) and ',' in item and '.' in item.replace(',', ''):
                                    result.append(float(item.replace(',', '')))
                                elif isinstance(item, str) and ',' in item:
                                    result.append(int(item.replace(',', '')))
                                elif isinstance(item, str) and item.replace('.', '', 1).isdigit() and item.count('.') == 1:
                                    result.append(float(item))
                                elif isinstance(item, str) and item.isdigit():
                                    result.append(int(item))
                                else:
                                    result.append(item)

    
                            last_row = len(sheet.col_values(1))
                            next_row = int(last_row) + 1
                            range_to_write = f'A{next_row}:S{next_row}'
                            sheet.update([result], range_to_write)
                    today_tday += timedelta(days=1)

                # 완료 로그
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtNaver.text()}-{brand}<br><span style='color:blue;'>완료</span>")

                time.sleep(2)

            except Exception as e:
                print(f"Error occurred: {e}")
                if edge_driver:
                    try:
                        edge_driver.close()  # 오류 발생 시 드라이버를 닫음
                        print("Driver closed successfully.")
                    except Exception as close_error:
                        print(f"Error closing driver: {close_error}")

                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtNaver.text()}-{brand}<br><span style='color:red;'>실패</span>")

        # def naveradInput(url, name, brand):

        if self.know_advtNaver.isChecked() == True:

            sheet_url = 'https://docs.google.com/spreadsheets/d/1CT15kvW9-ZLCJZNXrSsAe07eY9HLH2NTDRaDCqQU1h8/edit?gid=1582789201#gid=1582789201'
            sheet_name = '노마셀 네이버 R'
            target_url = "https://manage.searchad.naver.com/customers/2957190/reports/rtt-a001-000000000651985"
            # target_url = "https://manage.searchad.naver.com/customers/3067603/reports/rtt-a001-000000000842393"
            brand = "노마셀"
            
            advt_naver(target_url, sheet_url, sheet_name, brand)

        if self.zq_advtNaver.isChecked() == True:

            sheet_url = 'https://docs.google.com/spreadsheets/d/1U4s9UbjElH1QUk4-GvtTvxmHvzpswl22S8DWWkoWG9w/edit?gid=928641371#gid=928641371'
            sheet_name = '제니크 네이버 R'
            target_url = "https://manage.searchad.naver.com/customers/3163563/reports/rtt-a001-000000000725619"
            # target_url = "https://manage.searchad.naver.com/customers/3067603/reports/rtt-a001-000000000842393"
            brand = "제니크"
            
            advt_naver(target_url, sheet_url, sheet_name, brand)


# 네이버 gfa
        def advt_gfa(url, sheet_url, sheet_name, brand):

            edge_driver = None
            try:

                # EdgeOptions 설정
                edge_options = webdriver.EdgeOptions()
                edge_options.use_chromium = True
                edge_options.add_argument("disable-gpu")
                edge_options.add_argument("no-sandbox")
                edge_options.add_argument(f"user-data-dir={self.edge_path_folder.text()}")
                edge_options.add_argument("--profile-directory=Default")

                # 현재 실행 파일 기준 디렉토리
                if getattr(sys, 'frozen', False):
                    # PyInstaller로 패키징된 실행파일이면 실행파일 위치 기준
                    base_path = sys._MEIPASS if hasattr(sys, '_MEIPASS') else os.path.dirname(sys.executable)
                else:
                    # 개발환경(py로 실행)에서는 현재 경로
                    base_path = os.path.dirname(os.path.abspath(__file__))

                # 드라이버 경로 지정 (같은 폴더에 넣었다고 가정)
                driver_path = os.path.join(base_path, "msedgedriver.exe")

                if not os.path.exists(driver_path):
                    raise FileNotFoundError(f"Edge 드라이버가 없습니다: {driver_path}")

                # 드라이버 실행
                edge_service = Service(driver_path)
                edge_driver = webdriver.Edge(service=edge_service, options=edge_options)

                edge_driver.get(url)

                # 데이터 옵션선택
                WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#content > div > div.panel_body.report > div:nth-child(2) > div > div > div.ad_title > div > div.inner_right > div > div > div > button"))).click()
                # 데이터 옵션선택(rawdata)
                WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "rawdata")]'))).click()
                # 다운로드 요청
                WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#content > div > div.panel_body.report > div:nth-child(2) > div > div > div.ad_title > div > div.inner_right > a > button"))).click()
                # 확인 클릭
                WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#content > div > div.modal_root.__pa_fade_in > div > div.ly_content > div.ly_footer.type_border > button.button.button_panel.type_blue"))).click()
                print("확인클릭")
                time.sleep(2.5)
                # 닫기 클릭
                WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#content > div > div.modal_root.__pa_fade_in > div > div.ly_content > div.ly_footer.type_border > button"))).click()
                print("닫기클릭")
                time.sleep(2.5)
                # 다운로드 요청 목록
                WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "다운로드 요청 목록")]'))).click()
                print("다운로드요청목록")


                credential_file = 'triple-nectar-412808-da4dac0cc16e.json'
                client = gspread.service_account(filename=credential_file)
                spreadsheet = client.open_by_url(sheet_url)
                sheet = spreadsheet.worksheet(sheet_name)
                
                today_tday_temp = today_tday


                # 다운로드
                try:
                    element = WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#content > div > div.modal_root.__pa_fade_in > div > div.ly_content > div.ly_body.type_padding > div.table_area.table_sticky.active.scroll_horizon > div > div > div > div > div.table_whole > div > table > tbody > tr:nth-child(1) > td:nth-child(8) > div > button")))
                    
                    elementText = edge_driver.find_element(By.CSS_SELECTOR, "#content > div > div.modal_root.__pa_fade_in > div > div.ly_content > div.ly_body.type_padding > div.table_area.table_sticky.active.scroll_horizon > div > div > div > div > div.table_whole > div > table > tbody > tr:nth-child(1) > td:nth-child(8) > div > button").text

                    print(elementText)
                    if elementText == "데이터 없음":
                        self.logBox.append(f"{self.advt_group_box.title()}-{self.advtGFA.text()}-{brand}<br><span style='color:red;'>실패(데이터 없음)</span>")

                        WebDriverWait(edge_driver,5).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "전체 선택")]'))).click()

                        WebDriverWait(edge_driver,5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#content > div > div.modal_root.__pa_fade_in > div > div.ly_content > div.ly_body.type_padding > div.option_filter > div > div.inner_right > button"))).click()

                        WebDriverWait(edge_driver,5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#app > div.wrap > div.modal_root.__pa_fade_in > div > div.ly_content > div.ly_footer > button.button.button_panel.type_blue"))).click()

                        return
                    

                    # 다운로드 확인
                    cnt = 1
                    current_file_count1 = count_files(download_folder)
                    while cnt < 10:
                        element.click()
                        time.sleep(3)
                        current_file_count2 = count_files(download_folder)
                        if current_file_count1 != current_file_count2:
                            break
                        elif cnt == 300:
                            break

                        cnt += 1

                    time.sleep(2)
                    print("다운로드 완료")


                    get_latest_file_path = get_latest_file(download_folder)
                    print(get_latest_file_path)

                    current_file_count1 = count_files(download_folder)
                    print(current_file_count1)

                    def safe_extract(zip_path, extract_to):
                        """
                        ZIP 파일을 압축 해제하면서 파일 이름 충돌 시 다른 이름으로 저장
                        """
                        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                            for member in zip_ref.namelist():
                                original_path = os.path.join(extract_to, member)
                                target_path = original_path

                                # 파일 이름 충돌 시 새 이름 생성
                                if os.path.exists(original_path):
                                    base, ext = os.path.splitext(member)
                                    count = 1
                                    while os.path.exists(target_path):
                                        target_path = os.path.join(extract_to, f"{base}_{count}{ext}")
                                        count += 1

                                # 해당 파일 추출
                                with zip_ref.open(member) as source, open(target_path, 'wb') as target:
                                    target.write(source.read())

                    # 예제 사용
                    zip_path = get_latest_file_path  # 압축 파일 경로
                    extract_to = download_folder  # 압축 해제 경로

                    if not os.path.exists(extract_to):
                        os.makedirs(extract_to)

                    safe_extract(zip_path, extract_to)

                    current_file_count2 = count_files(download_folder)
                    print(current_file_count2)
                    time.sleep(1)
                    get_latest_file_path = get_latest_file(download_folder)
                    print(get_latest_file_path)

                    # 인코딩 감지
                    with open(get_latest_file_path, 'rb') as f:
                        result = chardet.detect(f.read())
                        detected_encoding = result['encoding']


                    # # 감지된 인코딩으로 파일을 읽기
                    # with open(get_latest_file_path, 'r', encoding=detected_encoding) as f:
                    #     reader = csv.reader(f)
                    #     rows = list(reader)
                    #     print("내용 읽기")
                    #     for row in rows:
                    #         print(row)


                    # csv 파일 변수 지정
                    csv_file = get_latest_file(download_folder)

                    
                    while today_tday_temp != today:
                        print(today_tday_temp)

                        result = []
                        today_tday_gfa = str(today_tday_temp).replace("-", ".")+ "."

                        # 타겟 날짜에 맞는 데이터 고르기
                        with open(csv_file, newline='', encoding=detected_encoding) as csvfile:
                            reader = csv.reader(csvfile)

                            # 각 행 리스트화
                            rows = list(reader)
                            reader_date = []

                            # 날짜만 따로 저장
                            for row in rows:
                                reader_date.append(row[2])
                            print(reader_date)

                            # 시트에 입력할 행 저장
                            last_row = len(sheet.col_values(1))
                            next_row = int(last_row) + 1

                            print(today_tday_gfa)

                            # 중복 입력 방지
                            if str(today_tday_temp) in sheet.col_values(1):
                                today_tday_temp += timedelta(days=1)
                                continue

                            # 목표날짜가 날짜 리스트에 있으면
                            elif today_tday_gfa in reader_date:
                                row_num = reader_date.index(today_tday_gfa)

                                for item in rows[row_num]:
                                    if isinstance(item, str) and '%' in item:
                                        result.append(float(item.strip('%')) / 100)
                                    elif isinstance(item, str) and ',' in item:
                                        result.append(int(item.replace(',', '')))
                                    elif isinstance(item, str) and item.replace('.', '', 1).isdigit() and item.count('.') == 1:
                                        result.append(float(item))
                                    elif isinstance(item, str) and item.isdigit():
                                        result.append(int(item))
                                    else:
                                        result.append(item)

                                range_to_write = f'B{next_row}:BY{next_row}'
                                sheet.update([result], range_to_write)
                                sheet.update([[str(today_tday_temp)]], f'A{next_row}')
                                continue

                            # 없으면 더미 데이터 입력    
                            else:
                                dumyData = ["-", "-", "-", 0, "-", 0, "-", 0, "-", "-", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
                                range_to_write = f'B{next_row}:BY{next_row}'
                                sheet.update([dumyData], range_to_write)
                                sheet.update([[str(today_tday_temp)]], f'A{next_row}')
                                continue

                        today_tday_temp += timedelta(days=1)

                except:

                    print("데이터 없음")

                    while today_tday_temp != today:
                        print(today_tday_temp)


                        # 시트에 입력할 행 저장
                        last_row = len(sheet.col_values(1))
                        next_row = int(last_row) + 1


                        # 중복 입력 방지
                        if str(today_tday_temp) in sheet.col_values(1):
                            today_tday_temp += timedelta(days=1)
                            continue

                        # 없으면 더미 데이터 입력    
                        else:
                            dumyData = ["-", "-", "-", 0, "-", 0, "-", 0, "-", "-", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
                            range_to_write = f'B{next_row}:BY{next_row}'
                            sheet.update([dumyData], range_to_write)
                            sheet.update([[str(today_tday_temp)]], f'A{next_row}')
                            continue

                    today_tday_temp += timedelta(days=1)
                
                WebDriverWait(edge_driver,5).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "전체 선택")]'))).click()

                WebDriverWait(edge_driver,5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#content > div > div.modal_root.__pa_fade_in > div > div.ly_content > div.ly_body.type_padding > div.option_filter > div > div.inner_right > button"))).click()

                WebDriverWait(edge_driver,5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#app > div.wrap > div.modal_root.__pa_fade_in > div > div.ly_content > div.ly_footer > button.button.button_panel.type_blue"))).click()


                try:
                        edge_driver.close()
                except (InvalidSessionIdException, WebDriverException) as e:
                        print("브라우저 세션이 이미 종료되었거나 무효함:", e)


                # 완료 로그
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtGFA.text()}-{brand}<br><span style='color:blue;'>완료</span>")

            except Exception as e:
                print(f"Error occurred: {e}")
                if edge_driver:
                    try:
                        edge_driver.close()  # 오류 발생 시 드라이버를 닫음
                        print("Driver closed successfully.")
                    except Exception as close_error:
                        print(f"Error closing driver: {close_error}")

                # 실패 로그
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtGFA.text()}-{brand}<br><span style='color:red;'>실패</span>")

        if self.know_advtGFA.isChecked() == True:
            url = "https://gfa.naver.com/adAccount/accounts/113881/report/performance?startDate=2025-08-22&endDate=2025-08-28&adUnit=AD_ACCOUNT&dateUnit=DAY&placeUnit=TOTAL&dimension=TOTAL&currentPage=1&pageSize=10&filterList=%5B%5D&showColList=%5B%22result%22,%22sales_per_result%22,%22sales%22,%22schedule%22,%22imp_count%22,%22cpm%22,%22click_count%22,%22cpc%22,%22ctr%22%5D&period=last7daysWithoutToday&accessAdAccountNo=113881"
            sheet_name = "노마셀 네이버 GFA R"
            brand = "노마셀"

            advt_gfa(url, sheet_advt_know_url, sheet_name, brand)

        if self.zq_advtGFA.isChecked() == True:
            url = "https://gfa.naver.com/adAccount/accounts/1615786/report/performance?startDate=2024-12-03&endDate=2024-12-09&adUnit=AD_ACCOUNT&dateUnit=DAY&placeUnit=TOTAL&dimension=TOTAL&currentPage=1&pageSize=10&filterList=%5B%5D&showColList=%5B%22result%22,%22sales_per_result%22,%22sales%22,%22schedule%22,%22imp_count%22,%22cpm%22,%22click_count%22,%22cpc%22,%22ctr%22%5D&period=last7daysWithoutToday&accessAdAccountNo=1615786"
            sheet_name = "제니크 GFA R"
            brand = "제니크"

            advt_gfa(url, sheet_advt_zenique_url, sheet_name, brand)

# 파워컨텐츠
         
        def advt_pc(url, url2,  id, pw, sheetUrl, sheetName, key, key2, brand):

            driver = None
            try:

                # 서비스 계정 키 파일 경로
                credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                # gspread 클라이언트 초기화
                client = gspread.service_account(filename=credential_file)

                # Google 시트 열기
                spreadsheet = client.open_by_url(sheetUrl)

                # 첫 번째 시트 선택
                sheet = spreadsheet.worksheet(sheetName)

                # 크롬 On
                ### chromedriver_autoinstaller.install() 사용 추가
                chromedriver_path = chromedriver_autoinstaller.install()
                chrome_options = webdriver.ChromeOptions()
                chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
                chrome_options.add_argument("--start-maximized") #최대 크기로 시작
                # chrome_options.add_argument('--incognito')
                # chrome_options.add_argument('--window-size=1920,1080')  
                # chrome_options.add_argument('--headless')
                chrome_options.add_experimental_option('detach', True)

                user_data = self.chrome_path_folder.text()
                chrome_options.add_argument(f"user-data-dir={user_data}")
                chrome_options.add_argument("--profile-directory=Profile 1")
                
                user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
                headers = {'user-agent' : user_agent}

                driver = webdriver.Chrome(
                    service=Service(chromedriver_path),
                    options=chrome_options
                )

                driver.get(url)

                ##################################### 로그인
                ##################################### 로그인
                ##################################### 로그인
                ##################################### 로그인

                ### 로그인
                # ID
                input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#mall_id")))
                input_field.click()
                time.sleep(1)
                input_field.send_keys(Keys.CONTROL + "a")
                input_field.send_keys(Keys.BACKSPACE)
                driver.find_element(By.CSS_SELECTOR, "#mall_id").send_keys(id)

                # PW
                input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#userpasswd")))
                input_field.click()
                input_field.send_keys(Keys.CONTROL + "a")
                input_field.send_keys(Keys.BACKSPACE)
                driver.find_element(By.CSS_SELECTOR, "#userpasswd").send_keys(pw)

                # 로그인클릭
                driver.find_element(By.CSS_SELECTOR,'#frm_user > div > div.mButton > button').click()

                #비밀번호변경안내
                try: WebDriverWait(driver, 5).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#iptBtnEm")))).click() 
                except: pass

                #화면로딩대기
                WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "오늘의 할 일")]')))

        ### 데이터 화면 접근
                driver.get(url2)
                WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#mCSB_2_container > ul:nth-child(1) > li:nth-child(10)"))).click() #통계 클릭
                WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "접속통계")]'))).click() #접속통계클릭

                #새 창 대기
                current_window_handle = driver.current_window_handle

                new_window_handle = None
                while not new_window_handle:
                    for handle in driver.window_handles:
                        if handle != current_window_handle:
                            new_window_handle = handle
                            break


                #팝업으로 제어 변경
                driver.switch_to.window(driver.window_handles[1]) 

                WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_pth")))).click() #방문경로분석
                WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_pth_dod")))).click() #방문도메인(상세)

        ### 데이터 검색
                # 어제 클릭
                WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(3) > a:nth-child(2) > img")))).click()

                updates = []
                formats = []
                dayUpdates = []

                for ii in range(target_days_input, 0, -1):

                    days_vis = datetime.timedelta(days=ii)
                    before_day_vis = today-days_vis

                    print(str(before_day_vis))
                    if str(before_day_vis) in sheet.col_values(1):
                        print("패스")
                        continue

                    # 달력클릭
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(2) > a:nth-child(3)")))).click()

                    #시작
                    #년도 선택
                    before_year = (before_day_vis).strftime("%Y")
                    select_element = driver.find_element(By.CSS_SELECTOR, '#yearSB_1')
                    select = Select(select_element)
                    select.select_by_value(before_year)

                    #달 선택
                    before_month = str(int((before_day_vis).strftime("%m")))
                    select_element = driver.find_element(By.CSS_SELECTOR, '#monthSB_1')
                    select = Select(select_element)
                    select.select_by_value(before_month)

                    #일 선택
                    before_day1 = str(int((before_day_vis).strftime("%d")))
                    for i in range(1, 43):
                        try:
                            element = driver.find_element(By.ID, f'li_{i}')
                            if element.text == before_day1:
                                element.click()
                                print("before_day1 clicked")
                                break

                        except:
                            print(f'li_{i} not found')


                    #끝
                    #년도 선택
                    select_element = driver.find_element(By.CSS_SELECTOR, '#yearSB_2')
                    select = Select(select_element)
                    select.select_by_value(before_year)

                    #달 선택
                    select_element = driver.find_element(By.CSS_SELECTOR, '#monthSB_2')
                    select = Select(select_element)
                    select.select_by_value(before_month)

                    #일 선택
                    for i in range(1, 43):
                        try:
                            element = driver.find_element(By.ID, f'le_{i}')
                            if element.text == before_day1:
                                element.click()
                                print("before_day1 clicked")
                                break

                        except:
                            print(f'le_{i} not found')

                    keywords = [key, key2]
                    # 검색어 입력(NV, NPO, GS)
                    for item in keywords:
                        print(item, "검색")
                        
                        search = driver.find_element(By.CSS_SELECTOR, "#body_center > table:nth-child(13) > tbody > tr > td:nth-child(1) > input")
                        search.send_keys(Keys.CONTROL + "a")
                        search.send_keys(Keys.BACKSPACE)
                        search.click()
                        search.send_keys(item)
                        
                        # 조회
                        element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR, "#body_center > table:nth-child(13) > tbody > tr > td:nth-child(1) > a > img"))))
                        element.click()
                        
                        last_row = len(sheet.get_all_values())
                        print(last_row)
                        next_row = last_row + 1  # 다음 행 번호

                        # 데이터 불러오기
                        for i in range(1,30):
                            
                            try:
                                
                                line = driver.find_element(By.CSS_SELECTOR, f"#detail_pfm_total > tr:nth-child({i})").text
                                lineSplit = line.strip().split(" ")
                                print(lineSplit)

                                

                                def convert_data(data):
                                    result = []
                                    for item in data:
                                        if isinstance(item, str) and '%' in item:
                                            result.append(float(item.strip('%')) / 100)
                                        elif isinstance(item, str) and ',' in item:
                                            result.append(int(item.replace(',', '')))
                                        elif item.isdigit():
                                            result.append(int(item))
                                        else:
                                            result.append(item)
                                    return result

                                # 입력할 데이터
                                converted_data = convert_data(lineSplit)

                                # 구글 시트에서 퍼센트 형식으로 변경하기 위해 셀 범위를 지정
                                # 예시에서는 C1, E1 셀을 퍼센트 형식으로 설정

                                range_to_write_day = f"A{next_row}"
                                    
                                # batch로 입력할 날짜 모으기
                                dayUpdates.append({'range': range_to_write_day, 'values': [[str(before_day_vis)]]})

                                # data 입력 범위
                                range_to_write = f'B{next_row}:I{next_row}'

                                # batch로 입력할 data 모으기
                                updates.append({'range': range_to_write, 'values': [converted_data]})

                                print("OK")

                                # 정렬할 format 세팅
                                formats.append({
                                'range': f"D{next_row}",
                                'format': {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}}
                                })
                                formats.append({
                                'range': f"F{next_row}",
                                'format': {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}}
                                })

                                next_row += 1

                            except:
                                if i == 1:
                                    dummyData = [item, '0', '0', '0', '0', '0', '0', '0']

                                    last_row = len(sheet.get_all_values())
                                    print(last_row)
                                    next_row = last_row + 1  # 다음 행 번호

                                    def convert_data(data):
                                        result = []
                                        for item in data:
                                            if isinstance(item, str) and '%' in item:
                                                result.append(float(item.strip('%')) / 100)
                                            elif isinstance(item, str) and ',' in item:
                                                result.append(int(item.replace(',', '')))
                                            elif item.isdigit():
                                                result.append(int(item))
                                            else:
                                                result.append(item)
                                        return result

                                    # 입력할 데이터
                                    converted_data = convert_data(dummyData)

                                    # 구글 시트에서 퍼센트 형식으로 변경하기 위해 셀 범위를 지정
                                    # 예시에서는 C1, E1 셀을 퍼센트 형식으로 설정
                                    sheet.format(f"D{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                                    sheet.format(f"F{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                                        
                                    sheet.update([[str(before_day_vis)]], f"A{next_row}")
                                    range_to_write = f'B{next_row}:I{next_row}'
                                    sheet.update([converted_data], range_to_write)
                                    print("OK")
                                    break

                            if updates:
                                print(updates)
                                print(dayUpdates)
                                sheet.batch_update(updates)
                                sheet.batch_update(dayUpdates)
                                for fmt in formats:
                                    sheet.format(fmt['range'], fmt['format'])
                                updates.clear()
                                dayUpdates.clear()
                                formats.clear()
                                time.sleep(0.5)  # 각 배치 요청 사이에 지연 시간을 추가
                                print("Batch update and format applied.")

                driver.close()
                time.sleep(0.1)
                
                driver.switch_to.window(driver.window_handles[0])
                driver.close()

                # 완료 로그
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtPC.text()}-{brand}<br><span style='color:blue;'>완료</span>")

            except Exception as e:
                print(f"Error occurred: {e}")
                if driver:
                    try:
                        for handle in driver.window_handles:
                            driver.switch_to.window(handle)  # 각 창으로 전환
                            driver.close()  # 현재 창 닫기
                    except Exception as close_error:
                        print(f"Error closing driver: {close_error}")

                # 실패 로그
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtPC.text()}-{brand}<br><span style='color:red;'>실패</span>")

        url_cafe24 = "https://eclogin.cafe24.com/Shop/" 
        
        if self.know_advtPC.isChecked() == True:
            cafe24_id_know = self.login_info("CAFE_KNOW_ID")
            cafe24_pw_know = self.login_info("CAFE_KNOW_PW")
            url2 = "https://fkark12.cafe24.com/disp/admin/shop1/report/DailyList"

            sheetName_knowPCR = "노마셀 파워콘텐츠 R"
            Keyword = "NV"
            Keyword2 = "GS"
            brand = "노마셀"

            advt_pc(url_cafe24, url2, cafe24_id_know, cafe24_pw_know, sheet_advt_know_url, sheetName_knowPCR, Keyword, Keyword2, brand)


# 구글 광고
         
        def advt_google(url_google):

            driver = None
            # try:

            # 크롬 On
            ### chromedriver_autoinstaller.install() 사용 추가
            chromedriver_path = chromedriver_autoinstaller.install()
            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_argument("--start-maximized") #최대 크기로 시작
            # chrome_options.add_argument('--incognito')
            # chrome_options.add_argument('--window-size=1920,1080')  
            # chrome_options.add_argument('--headless')
            chrome_options.add_experimental_option('detach', True)

            user_data = self.chrome_path_folder.text()
            chrome_options.add_argument(f"user-data-dir={user_data}")
            chrome_options.add_argument("--profile-directory=Profile 1")
            
            user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
            headers = {'user-agent' : user_agent}

            driver = webdriver.Chrome(
                service=Service(chromedriver_path),
                options=chrome_options
            )

            driver.get(url_google)
            try:
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.CLASS_NAME, 'button-text')))
            except:
                driver.refresh()
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.CLASS_NAME, 'button-text')))

            calOpen = driver.find_element(By.CLASS_NAME, 'button-text')

            time.sleep(1)
            # 달력 열기
            calOpen.click()

            # 날짜 선택
            WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CLASS_NAME, 'visible-month')))
            time.sleep(1.5)
            elements = driver.find_elements(By.CSS_SELECTOR, 'material-select-item')
            for element in elements:
                if '지난 7일(어제까지)' in element.text:
                    print("Element found:", element.text)
                    element.click()
                    break
            # WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "지난 7일(어제까지)")]')))
            # weekElement = driver.find_element(By.XPATH, '//*[contains(text(), "지난 7일(어제까지)")]')
            # weekElement.click()

            time.sleep(1)
            # 11. 구글 다운로드 클릭, CSS선택자 방식에서 XPATH - contains, text() 방식으로 변경(CSS선택자 매번바뀜)
            schedule = driver.find_element(By.XPATH, "//*[contains(text(), '일정')]")
            #다운
            # 부모의 부모의 부모의 부모의 부모의 이전 요소 찾기 및 클릭
            previous_sibling = schedule.find_element(By.XPATH, "ancestor::*[4]/preceding-sibling::*[1]")
            previous_sibling.click()


            # Excel .csv 선택
            lists = driver.find_element(By.CLASS_NAME, "download-dropdown").find_elements(By.CLASS_NAME, "item")

            for item in lists:
                if item.text == "Excel .csv":
                    item.click()
                    break

            ## 구글 다운로드 실패 시 재시도 적용
            try:
                # 다운로드 확인
                cnt = 1
                current_file_count1 = count_files(download_folder)
                while cnt < 10:

                    # 11. 구글 다운로드 클릭, CSS선택자 방식에서 XPATH - contains, text() 방식으로 변경(CSS선택자 매번바뀜)
                    schedule = driver.find_element(By.XPATH, "//*[contains(text(), '일정')]")
                    #다운
                    # 부모의 부모의 부모의 부모의 부모의 이전 요소 찾기 및 클릭
                    previous_sibling = schedule.find_element(By.XPATH, "ancestor::*[4]/preceding-sibling::*[1]")
                    previous_sibling.click()


                    # Excel .csv 선택
                    lists = driver.find_element(By.CLASS_NAME, "download-dropdown").find_elements(By.CLASS_NAME, "item")
                    for item in lists:
                        if item.text == "Excel .csv":
                            item.click()
                            break
                    time.sleep(3)
                    current_file_count2 = count_files(download_folder)
                    if current_file_count1 != current_file_count2:
                        break
                    elif cnt == 30:
                        break

                    cnt += 1

            except:


                # 다운로드 확인
                cnt = 1
                current_file_count1 = count_files(download_folder)
                while cnt < 10:
                    for item in lists:
                        # 11. 구글 다운로드 클릭, CSS선택자 방식에서 XPATH - contains, text() 방식으로 변경(CSS선택자 매번바뀜)
                        schedule = driver.find_element(By.XPATH, "//*[contains(text(), '일정')]")
                        #다운
                        # 부모의 부모의 부모의 부모의 부모의 이전 요소 찾기 및 클릭
                        previous_sibling = schedule.find_element(By.XPATH, "ancestor::*[4]/preceding-sibling::*[1]")
                        previous_sibling.click()


                        # Excel .csv 선택
                        lists = driver.find_element(By.CLASS_NAME, "download-dropdown").find_elements(By.CLASS_NAME, "item")

                        if item.text == "Excel .csv":
                            item.click()
                            break
                    time.sleep(3)
                    current_file_count2 = count_files(download_folder)
                    if current_file_count1 != current_file_count2:
                        break
                    elif cnt == 30:
                        break

                    cnt += 1


## 구글 다운로드 실패 시 재시도 적용
            try:
                check_download()
            except:
                schedule = driver.find_element(By.XPATH, "//*[contains(text(), '일정')]")
                #다운
                # 부모의 부모의 부모의 부모의 부모의 이전 요소 찾기 및 클릭
                previous_sibling = schedule.find_element(By.XPATH, "ancestor::*[4]/preceding-sibling::*[1]")
                previous_sibling.click()


                # Excel .csv 선택
                lists = driver.find_element(By.CLASS_NAME, "download-dropdown").find_elements(By.CLASS_NAME, "item")

                for item in lists:
                    print(item.text)
                    if item.text == "Excel .csv":
                        item.click()
                        break
                check_download()


            time.sleep(2)
            driver.close()

            # except Exception as e:
            #     print(f"Error occurred: {e}")
            #     if driver:
            #         try:
            #             driver.close()  # 오류 발생 시 드라이버를 닫음
            #             print("Driver closed successfully.")
            #         except Exception as close_error:
            #             print(f"Error closing driver: {close_error}")
         
        def advt_google_rawdata(sheet_url, sheet_name, brand):

            target_days = target_days_input
            dayx = timedelta(days=target_days)
            day1 = timedelta(days=1)

            # 오늘 날짜 구하기
            today_yday = today-day1
            today_tday = today-dayx
            
            csv_file = get_latest_file(download_folder)

            # 서비스 계정 키 파일 경로
            credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

            # gspread 클라이언트 초기화
            client = gspread.service_account(filename=credential_file)

            # Google 시트 열기
            spreadsheet = client.open_by_url(sheet_url)

            # 첫 번째 시트 선택
            sheet = spreadsheet.worksheet(sheet_name)

            selected_rows = []

            with open(csv_file, newline='', encoding='utf-16') as csvfile:
                reader = csv.reader(csvfile)
                for i, row in enumerate(reader):
                    if 3 <= i <= 100:  # 범위 내 행 적용

                        # 데이터를 올바르게 파싱하기 위해 먼저 전체 문자열을 하나로 합친다
                        full_data = "".join(row)

                        # 탭(\t)으로 데이터를 분리한다
                        parsed_data = full_data.split('\t')

                        cleaned_data = [item.replace('"', '') for item in parsed_data]
                        selected_rows.append(cleaned_data)
                print(selected_rows)

            updates = []
            formats = []

            while today_tday != today:

                last_row = len(sheet.col_values(1))
                next_row = int(last_row) + 1
                print(next_row)

                print(today_tday, "검색 시작")
                for i in selected_rows:
                    if str(today_tday) in i:
                        new_selected_rows = []
                        result = []
                        print(today_tday, "찾음!")

                        for item in i:
                            if isinstance(item, str) and '%' in item:
                                result.append(float(item.strip('%')) / 100)
                            elif isinstance(item, str) and ',' in item and '.' in item.replace(',', ''):
                                result.append(float(item.replace(',', '')))
                            elif isinstance(item, str) and ',' in item:
                                result.append(int(item.replace(',', '')))
                            elif isinstance(item, str) and item.replace('.', '', 1).isdigit() and item.count('.') == 1:
                                result.append(float(item))
                            elif isinstance(item, str) and item.isdigit():
                                result.append(int(item))
                            else:
                                result.append(item)

                        new_selected_rows.append(result)
                        print(new_selected_rows)

                        if brand == "노마셀":
                            print("노마셀 입력 시작")
                            print(new_selected_rows)

                            range_to_write = f'A{next_row}:K{next_row}'
                            updates.append({'range': range_to_write, 'values': new_selected_rows})

                            sheet.update(new_selected_rows, range_to_write)

                            # 셀 포맷 설정
                            sheet.format(f"G{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                            sheet.format(f'H{next_row}:J{next_row}', {"numberFormat": {"type":'NUMBER'}})
                            time.sleep(1)

                        next_row += 1

                today_tday += timedelta(days=1)
            
                # 완료 로그
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtGgle.text()}-{brand}<br><span style='color:blue;'>완료</span>")


        if self.know_advtGgle.isChecked() == True:
            url_ads_know = 'https://ads.google.com/aw/reporteditor/view?ocid=1379143590&workspaceId=-1615213561&reportId=928192574&euid=1114690018&__u=8943315282&uscid=1379143590&__c=4267857910&authuser=0'
            sheet_name_goog = "노마셀 구글 R"
            brand = "노마셀"

            try:
                advt_google(url_ads_know)
                advt_google_rawdata(sheet_advt_know_url, sheet_name_goog, brand)
            except Exception as e:
                print(f"Error occurred in advt_Google: {e}")
                # 실패 로그
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtGgle.text()}-{brand}<br><span style='color:red;'>실패</span>")


# 메타 광고
        def meta_rawdata(sheet_url, sheet_name, know_TF, brand):

            xlsx_file = get_latest_file(download_folder)
            wb = load_workbook(xlsx_file)
            ws = wb.active

            # 서비스 계정 키 파일 경로
            credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

            # gspread 클라이언트 초기화
            client = gspread.service_account(filename=credential_file)

            # Google 시트 열기
            spreadsheet = client.open_by_url(sheet_url)

            # 첫 번째 시트 선택
            sheet = spreadsheet.worksheet(sheet_name)

            # else:
            data_to_paste = []
            data_to_pasteDay = []
            today_tdayTemp = today_tday

            # 8. 메타 n일전 데이터 불러오기(데이터 없으면 더미데이터 입력)
            # 두 번째 행이 비어있는지 확인
            second_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
            if second_row and all(cell is None for cell in second_row[0]):
                while today_tdayTemp != today:
                    metaDataEmpty = [str(today_tdayTemp), '-', 0, 0, 0, 0, 0, 0, 0, 0, '-', 0, 0, 0, 0, str(today_tdayTemp), str(today_tdayTemp)]

                    # 중복 입력 방지
                    if str(today_tdayTemp) in sheet.col_values(1):
                        today_tdayTemp += timedelta(days=1)
                        continue

                    last_row = len(sheet.col_values(1))
                    print(last_row)
                    next_row = last_row + 1  # 다음 행 번호

                    # 데이터 추가
                    range_to_write = f'A{next_row}:Q{next_row}'
                    sheet.update([metaDataEmpty], range_to_write) #한줄

                    today_tdayTemp += timedelta(days=1)
                    

            else:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    changed_row = list(row)
                    changed_row[0], changed_row[1] = changed_row[1], changed_row[0]
                    data_to_paste.append(changed_row)
                data_to_paste.reverse()
                print(data_to_paste)
                for i in data_to_paste:
                    data_to_pasteDay.append(i[0])

                print(data_to_pasteDay)

                print(today_tdayTemp)
                print(today_yday)

                while today_tdayTemp != today:

                    last_row = len(sheet.col_values(1))
                    print(last_row)
                    next_row = last_row + 1  # 다음 행 번호

                    # 중복 입력 방지
                    if str(today_tdayTemp) in sheet.col_values(1):
                        today_tdayTemp += timedelta(days=1)
                        continue

                    # 해당 날짜 데이터 있으면
                    if str(today_tdayTemp) in data_to_pasteDay:
                        for num, i in enumerate(data_to_pasteDay):
                            if i == str(today_tdayTemp):

                                data_to_paste[num]

                                # 데이터 추가
                                range_to_write = f'A{next_row}:Q{next_row}'
                                sheet.update([data_to_paste[num]], range_to_write) #한줄

                                next_row += 1

                    # 해당 날짜 데이터 없으면 더미데이터
                    else:
                        metaDataEmpty = [str(today_tdayTemp), '-', 0, 0, 0, 0, 0, 0, 0, 0, '-', 0, 0, 0, 0, str(today_tdayTemp), str(today_tdayTemp)]

                        # 데이터 추가
                        range_to_write = f'A{next_row}:Q{next_row}'
                        sheet.update([metaDataEmpty], range_to_write) #한줄

                    today_tdayTemp += timedelta(days=1)

            # 완료 로그
            self.logBox.append(f"{self.advt_group_box.title()}-{self.advtMeta.text()}-{brand}<br><span style='color:blue;'>완료</span>")

# 메타
        def advt_meta(url_meta, know_TF):

            driver = None
            try:

                # 크롬 On
                chromedriver_path = chromedriver_autoinstaller.install()
                chrome_options = webdriver.ChromeOptions()
                chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
                chrome_options.add_argument("--start-maximized") #최대 크기로 시작
                chrome_options.add_experimental_option('detach', True)

                user_data = self.chrome_path_folder.text()
                chrome_options.add_argument(f"user-data-dir={user_data}")
                chrome_options.add_argument("--profile-directory=Profile 1")
                
                user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
                headers = {'user-agent' : user_agent}

                driver = webdriver.Chrome(
                    service=Service(chromedriver_path),
                    options=chrome_options
                )

                driver.get(url_meta)

                # meta_id = 'healer10@kakao.com'
                # meta_pw = 'fhdifxmfl1305!!'

                # '비밀번호를' 텍스트를 포함하는 요소 찾기
                time.sleep(2)

                try:
                    WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "비밀번호를")]')))
                    pw = driver.find_element(By.XPATH, '//*[contains(text(), "비밀번호를")]')

                    if pw:
                        print("pw 만족")
                        # 이전 형제 요소 찾기
                        parent_element = pw.find_element(By.XPATH, '..')
                        previous_sibling = parent_element.find_element(By.XPATH, 'preceding-sibling::*[1]')
                        print("Previous sibling found:", previous_sibling.text)
                        print
                        previous_sibling.click()
                    
                    else:
                        print("요소를 찾을 수 없습니다.")

                    time.sleep(1)
                    driver.get(url_meta)

                except:
                    pass
                #알림 제거
                try:
                    body = driver.find_element(By.CSS_SELECTOR, 'body')
                    ActionChains(driver).move_to_element(body).click().perform()
                except: pass


                # 달력 열기
                WebDriverWait(driver, 300).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="PNG_EXPORT"]/div/div[3]/div[1]/div[1]/div/div/div/div[1]/div[2]/div[2]/div/div/div/span')))
                driver.find_element(By.XPATH, '//*[@id="PNG_EXPORT"]/div/div[3]/div[1]/div[1]/div/div/div/div[1]/div[2]/div[2]/div/div/div/span').click()

                # 오늘 선택하기
                WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "최근 7일")]'))).click()

                element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), '내보내기')]")))

                

                
                # 다운로드 확인
                cnt = 1
                current_file_count1 = count_files(download_folder)
                while cnt < 10:
                    element.click() #내보내기
                    time.sleep(1.5)

                    # 11. 구글 다운로드 클릭, CSS선택자 방식에서 XPATH - contains, text() 방식으로 변경(CSS선택자 매번바뀜)
                    cancelElement = driver.find_element(By.XPATH, "//*[contains(text(), '취소')]")
                    #다운
                    # 부모의 부모의 부모의 부모의 부모의 이전 요소 찾기 및 클릭
                    previous_sibling = cancelElement.find_element(By.XPATH, "ancestor::*[4]/following-sibling::*[1]")
                    previous_sibling.click()
                    
                    time.sleep(5)
                    current_file_count2 = count_files(download_folder)
                    if current_file_count1 != current_file_count2:
                        break

                    cnt += 1

                    
                time.sleep(3)

                driver.close()

            except Exception as e:
                print(f"Error occurred: {e}")
                if driver:
                    try:
                        driver.close()  # 오류 발생 시 드라이버를 닫음
                        print("Driver closed successfully.")
                    except Exception as close_error:
                        print(f"Error closing driver: {close_error}")

        #메타 천명연구소
        if self.CMlabs_advtMeta.isChecked() == True:
            url_meta_CMlabs = 'https://adsmanager.facebook.com/adsmanager/reporting/view?act=1314300116713919&business_id=635001998695042&selected_report_id=120233079125660077' #노마셀
            sheet_name_CMlabs = '천명연구소 페이스북 R'
            know_TF = 1
            brand = "천명연구소"

            # try:
            advt_meta(url_meta_CMlabs, know_TF)
            meta_rawdata(sheet_advt_CMlabs_url, sheet_name_CMlabs, know_TF, brand)
            # except Exception as e:
            #     print(f"Error occurred in advt_Meta: {e}")
            #     # 실패 로그
            #     self.logBox.append(f"{self.advt_group_box.title()}-{self.advtMeta.text()}-{brand}<br><span style='color:red;'>실패</span>")
        
        #메타 노마셀
        if self.know_advtMeta.isChecked() == True:
            # url_meta_knowmycell = 'https://adsmanager.facebook.com/adsmanager/reporting/view?act=238068255778220&business_id=635001998695042&selected_report_id=120200841324100083' #노마셀
            url_meta_knowmycell = 'https://adsmanager.facebook.com/adsmanager/reporting/view?act=1362104928634334&ads_manager_write_regions=true&business_id=635001998695042&selected_report_id=120235417435240087' #노마셀
            
            sheet_name_know = '노마셀 페이스북 R'
            know_TF = 1
            brand = "노마셀"

            try:
                advt_meta(url_meta_knowmycell, know_TF)
                meta_rawdata(sheet_advt_know_url, sheet_name_know, know_TF, brand)
            except Exception as e:
                print(f"Error occurred in advt_Meta: {e}")
                # 실패 로그
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtMeta.text()}-{brand}<br><span style='color:red;'>실패</span>")

        #메타 제니크
        if self.zq_advtMeta.isChecked() == True:
            url_meta_zq = 'https://adsmanager.facebook.com/adsmanager/reporting/view?act=7003471889761390&business_id=635001998695042&selected_report_id=120211428882470776' #제니크
            sheet_name_zq = '제니크 페이스북 R'
            know_TF = 1
            brand = "제니크"

            try:
                advt_meta(url_meta_zq, know_TF)
                meta_rawdata(sheet_advt_zenique_url, sheet_name_zq, know_TF, brand)
            except Exception as e:
                print(f"Error occurred in advt_Meta: {e}")
                # 실패 로그
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtMeta.text()}-{brand}<br><span style='color:red;'>실패</span>")



        def advt_tiktok(url, sheet_url, sheet_name, brand):
            driver = None
            try:
                # 크롬 On
                ### chromedriver_autoinstaller.install() 사용 추가
                chromedriver_path = chromedriver_autoinstaller.install()
                chrome_options = webdriver.ChromeOptions()

                # [기존] 자동화 배너 제거
                chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
                chrome_options.add_experimental_option("useAutomationExtension", False)

                chrome_options.add_argument("--start-maximized")  # 최대 크기로 시작
                chrome_options.add_experimental_option('detach', True)

                # [선택] 언어/지역 설정 (ko-KR 우선)
                chrome_options.add_argument("--lang=ko-KR,ko")

                # [중요] UA는 headers로 넣어도 네트워크 요청에는 반영 안됩니다.
                # 반드시 옵션/Network.setUserAgentOverride로 설정하세요.
                user_agent = (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/117.0.0.0 Safari/537.36"
                )
                chrome_options.add_argument(f"--user-agent={user_agent}")

                # [기존 프로필 유지]
                user_data = self.chrome_path_folder.text()
                chrome_options.add_argument(f"user-data-dir={user_data}")
                chrome_options.add_argument("--profile-directory=Profile 1")

                # [권장] 자동화 탐지 플래그 비활성화
                chrome_options.add_argument("--disable-blink-features=AutomationControlled")

                driver = webdriver.Chrome(
                    service=Service(chromedriver_path),
                    options=chrome_options
                )

                # ====== 🔹 Stealth 유사 패치 (Puppeteer StealthPlugin 대체) ======
                # (의도적으로 media.codecs / iframe.contentWindow 패치는 '미적용')
                # - navigator.webdriver 제거
                # - languages, language, platform, deviceMemory, hardwareConcurrency 스푸핑
                # - plugins 존재하도록 스푸핑
                # - permissions.query(Notifications) 오버라이드
                # - WebGL vendor/renderer 스푸핑
                # - window.chrome.runtime 존재하도록 보강
                # - UA/Platform을 CDP로도 오버라이드 (일부 사이트는 CDP 경로만 신뢰)

                # 1) UA/플랫폼 CDP 오버라이드
                driver.execute_cdp_cmd("Network.enable", {})
                driver.execute_cdp_cmd(
                    "Network.setUserAgentOverride",
                    {"userAgent": user_agent, "platform": "Windows"}
                )

                # 2) 페이지 로드 전에 스크립트 삽입
                driver.execute_cdp_cmd(
                    "Page.addScriptToEvaluateOnNewDocument",
                    {
                        "source": r"""
                // == Selenium Stealth-like Patches ==
                (() => {
                // 헬퍼: 읽기전용 getter 덮어쓰기
                const defineGetter = (obj, prop, val) => {
                    try {
                    Object.defineProperty(obj, prop, { get: () => val, configurable: true });
                    } catch(e) {}
                };

                // webdriver 제거
                defineGetter(navigator, 'webdriver', undefined);

                // 언어/플랫폼/디바이스 스푸핑
                defineGetter(navigator, 'language', 'ko-KR');
                defineGetter(navigator, 'languages', ['ko-KR', 'ko']);
                defineGetter(navigator, 'platform', 'Win32');
                defineGetter(navigator, 'hardwareConcurrency', 8);
                defineGetter(navigator, 'deviceMemory', 8);

                // plugins 존재하도록
                Object.defineProperty(navigator, 'plugins', {
                    get: () => [1, 2, 3, 4, 5],
                });

                // Chrome 객체/런타임 존재하도록
                window.chrome = window.chrome || {};
                if (!window.chrome.runtime) {
                    window.chrome.runtime = {};
                }

                // permissions.query 오버라이드 (알림 권한 탐지 회피)
                const originalQuery = window.navigator.permissions && window.navigator.permissions.query;
                if (originalQuery) {
                    window.navigator.permissions.query = (parameters) => {
                    if (parameters && parameters.name === 'notifications') {
                        return Promise.resolve({ state: Notification.permission });
                    }
                    return originalQuery(parameters);
                    };
                }

                // WebGL vendor/renderer 스푸핑
                const getParameter = WebGLRenderingContext.prototype.getParameter;
                WebGLRenderingContext.prototype.getParameter = function(param) {
                    // UNMASKED_VENDOR_WEBGL
                    if (param === 0x9245) return 'Intel Inc.';
                    // UNMASKED_RENDERER_WEBGL
                    if (param === 0x9246) return 'Intel Iris OpenGL Engine';
                    return getParameter.apply(this, arguments);
                };

                // ❌ 미적용: media.codecs (Puppeteer stealth에서 지운 것과 동일하게 here도 건드리지 않음)
                // ❌ 미적용: iframe.contentWindow 패치
                })();
                """
                    }
                )

                # 이제 접속
                driver.get(url)

                today_tdayTemp = today_tday

                # 서비스 계정 키 파일 경로
                credential_file = 'triple-nectar-412808-da4dac0cc16e.json'
                # gspread 클라이언트 초기화
                client = gspread.service_account(filename=credential_file)
                # Google 시트 열기
                spreadsheet = client.open_by_url(sheet_url)
                print(spreadsheet)
                # 첫 번째 시트 선택
                sheet = spreadsheet.worksheet(sheet_name)
                print(sheet)
                print("today_tdayTemp:", today_tdayTemp)
                time.sleep(4)

                while today_tdayTemp != today:
                    print("today_tdayTemp:", today_tdayTemp)


                    # 첫 번째 열의 모든 값 가져오기
                    column_values = sheet.col_values(1)
                    print(column_values)
                    # 이미 입력되어있는지 확인
                    if str(today_tdayTemp) in column_values:
                        today_tdayTemp += timedelta(days=1)
                        print("+1 day")
                        continue

                    # 시작 날짜 입력
                    input_field1 = WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[placeholder='시작일']")))
                    input_field1.send_keys(Keys.CONTROL + "a")
                    input_field1.send_keys(str(today_tdayTemp))
                    
                    # 종료 날짜 입력
                    input_field2 = WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[placeholder='종료일']")))
                    input_field2.send_keys(Keys.CONTROL + "a")
                    input_field2.send_keys(str(today_tdayTemp))
                    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"button.refresh-btn,button.vi-button.refresh-btn,button.vi-byted-button.refresh-btn,button[data-testid^='button-button-'].refresh-btn"))).click()

                    time.sleep(1)
                    xpath = "//div[@role='dialog'][.//label[normalize-space()='보고서 이름 지정']]//button[.//span[normalize-space()='내보내기']]"
                    try:
                        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[3]/div/div/div[2]/div[3]/section/div/div/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/button'))).click()
                        element = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                        element.click()
                    except:
                        cancel_element = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[13]/div/div[3]/div/button[1]")))
                        cancel_element.click()
                        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[3]/div/div/div[2]/div[3]/section/div/div/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/button'))).click()
                        element = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                        element.click()
                    
                    
                    # 다운로드 확인
                    cnt = 1
                    current_file_count1 = count_files(download_folder)
                    while cnt < 10:
                        time.sleep(2)
                        current_file_count2 = count_files(download_folder)
                        if current_file_count1 != current_file_count2:
                            break
                        elif cnt == 300:
                            break
                        try:
                            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[3]/div/div/div[2]/div[3]/section/div/div/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/button'))).click()
                            element = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[12]/div/div[3]/div/button[2]")))
                            element.click()
                        except:
                            pass

                        cnt += 1

                    time.sleep(2)
                    print("다운로드 완료")

                    last_row = len(column_values)
                    next_row = last_row + 1  # 다음 행 번호

                    xlsx_file = get_latest_file(download_folder)
                    # 엑셀 2번째 행
                    df = pd.read_excel(xlsx_file, dtype=object)   # dtype=object로 읽기(혼합형 보존)
                    raw = df.iloc[0].tolist()

                    def to_cell(v):
                        if pd.isna(v):
                            return ""
                        if isinstance(v, (pd.Timestamp, dt.datetime, dt.date)):
                            return v.strftime("%Y-%m-%d %H:%M:%S")  # 필요시 포맷 조정
                        if isinstance(v, (np.integer, np.floating, np.bool_)):
                            return v.item()                         # numpy 스칼라 → 파이썬 스칼라
                        return v

                    row2 = [to_cell(v) for v in raw]

                    # ✅ 1) A{next_row}: 날짜만 (values 먼저!)
                    sheet.update([[str(today_tdayTemp)]], f"A{next_row}", value_input_option="USER_ENTERED")

                    # ✅ 2) B{next_row}: 엑셀 2행 전체 (values 먼저!)
                    sheet.update([row2], f"B{next_row}", value_input_option="USER_ENTERED")



                    today_tdayTemp += timedelta(days=1)

                driver.close()

                # 완료 로그
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtTiktok.text()}-{brand}<br><span style='color:blue;'>완료</span>")

            except Exception as e:
                print(f"Error occurred: {e}")
                if driver:
                    try:
                        driver.close()  # 오류 발생 시 드라이버를 닫음
                        print("Driver closed successfully.")
                    except Exception as close_error:
                        print(f"Error closing driver: {close_error}")

                # 실패 로그
                self.logBox.append(f"{self.advt_group_box.title()}-{self.advtTiktok.text()}-{brand}<br><span style='color:red;'>실패</span>")
                



        #틱톡 제니크
        if self.zq_advtTiktok.isChecked() == True:
            url_tiktok_zq = 'https://ads.tiktok.com/i18n/reporting/pivot/table/edit?reportId=7494132392777482247&aadvid=7353112763541258256' #제니크
            sheet_name_zq = '제니크 틱톡 R'
            brand = "제니크"
            advt_tiktok(url_tiktok_zq, sheet_advt_zenique_url, sheet_name_zq, brand)

        #틱톡 노마셀
        if self.know_advtTiktok.isChecked() == True:
            url_tiktok_know = 'https://ads.tiktok.com/i18n/reporting/pivot/table/edit?reportId=7363549941156806673&aadvid=7361287611433975825' #노마셀
            sheet_name_know = '노마셀 틱톡 R'
            brand = "노마셀"
            advt_tiktok(url_tiktok_know, sheet_advt_know_url, sheet_name_know, brand)

        
# 방문자수
         
        def visitors(url, id, pw, sheet_url, sheet_name, brand):

            driver = None
            # try:

            # 크롬 On
            chromedriver_path = chromedriver_autoinstaller.install()
            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_argument("--start-maximized") #최대 크기로 시작
            chrome_options.add_experimental_option('detach', True)

            user_data = self.chrome_path_folder.text()
            chrome_options.add_argument(f"user-data-dir={user_data}")
            chrome_options.add_argument("--profile-directory=Profile 1")
        
            user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
            headers = {'user-agent' : user_agent}

            driver = webdriver.Chrome(
                service=Service(chromedriver_path),
                options=chrome_options
            )

            driver.get(url_cafe24)

            ##################################### 로그인
            ##################################### 로그인
            ##################################### 로그인
            ##################################### 로그인

            # ID
            input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#mall_id")))
            input_field.click()
            time.sleep(1)
            input_field.send_keys(Keys.CONTROL + "a")
            input_field.send_keys(Keys.BACKSPACE)
            driver.find_element(By.CSS_SELECTOR, "#mall_id").send_keys(id)

            # PW
            input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#userpasswd")))
            input_field.click()
            input_field.send_keys(Keys.CONTROL + "a")
            input_field.send_keys(Keys.BACKSPACE)
            driver.find_element(By.CSS_SELECTOR, "#userpasswd").send_keys(pw)

            # 로그인클릭
            driver.find_element(By.CSS_SELECTOR,'#frm_user > div > div.mButton > button').click()

            #비밀번호변경안내
            try: WebDriverWait(driver, 5).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#iptBtnEm")))).click() 
            except: pass

            try:
                time.sleep(3)
                popup = driver.find_element(By.XPATH, '//*[contains(text(), "오늘 하루 보지 않기")]')
                popup.click()

            except: pass

            #화면로딩대기
            WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "오늘의 할 일")]')))
            ###################################### 조회수
            ###################################### 조회수
            ###################################### 조회수
            ###################################### 조회수

            driver.find_element(By.CSS_SELECTOR, "#mCSB_2_container > ul:nth-child(1) > li:nth-child(9)").click() #통계 클릭
            WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, "//*[contains(text(), '접속통계')]"))).click() #접속통계클릭 #접속통계클릭

            #새 창 대기
            current_window_handle = driver.current_window_handle

            new_window_handle = None
            while not new_window_handle:
                for handle in driver.window_handles:
                    if handle != current_window_handle:
                        new_window_handle = handle
                        break


            #팝업으로 제어 변경
            driver.switch_to.window(driver.window_handles[1]) 

            WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_pth")))).click() #방문경로분석
            WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_pth_dod")))).click() #방문도메인(상세)
            WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(3) > a:nth-child(2) > img")))).click()

            def click_element_with_retry(driver, selector, retries=3):
                for attempt in range(retries):
                    try:
                        element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector)))
                        element.click()
                        return True  # 클릭 성공 시 함수 종료
                    except StaleElementReferenceException:
                        print(f"[Retry {attempt+1}/{retries}] StaleElementReferenceException 발생, 요소 다시 찾는 중...")
                
                print("[Error] 요소를 찾을 수 없습니다.")
                return False  # 최종적으로 실패

            for ii in range(1, target_days_input+1):

                # 달력클릭
                click_element_with_retry(driver, '#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(2) > a:nth-child(3)', 3)

                days_vis = datetime.timedelta(days=ii)
                before_day_vis = today-days_vis
                

                #시작
                #년도 선택
                before_year = (before_day_vis).strftime("%Y")
                select_element = driver.find_element(By.CSS_SELECTOR, '#yearSB_1')
                select = Select(select_element)
                select.select_by_value(before_year)

                #달 선택
                before_month = str(int((before_day_vis).strftime("%m")))
                select_element = driver.find_element(By.CSS_SELECTOR, '#monthSB_1')
                select = Select(select_element)
                select.select_by_value(before_month)

                #일 선택
                before_day1 = str(int((before_day_vis).strftime("%d")))
                for i in range(1, 43):
                    try:
                        element = driver.find_element(By.ID, f'li_{i}')
                        if element.text == before_day1:
                            element.click()
                            print("before_day1 clicked")
                            break

                    except:
                        print(f'li_{i} not found')


                #끝
                #년도 선택
                select_element = driver.find_element(By.CSS_SELECTOR, '#yearSB_2')
                select = Select(select_element)
                select.select_by_value(before_year)

                #달 선택
                select_element = driver.find_element(By.CSS_SELECTOR, '#monthSB_2')
                select = Select(select_element)
                select.select_by_value(before_month)

                #일 선택
                for i in range(1, 43):
                    try:
                        element = driver.find_element(By.ID, f'le_{i}')
                        if element.text == before_day1:
                            element.click()
                            print("before_day1 clicked")
                            break

                    except:
                        print(f'le_{i} not found')

                # 조회
                click_element_with_retry(driver, '#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(2) > a:nth-child(4) > img', 3)


                visitors = driver.find_elements(By.ID, "summary_pfm_total")
                for num in visitors:

                    for attempt in range(3):
                        try:
                            if WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#summary_pfm_total > td:nth-child(2)"))):
                                element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#summary_pfm_total > td:nth-child(2)")))
                                break
                            
                        except StaleElementReferenceException:
                            print(f"[Retry {attempt+1}/{3}] StaleElementReferenceException 발생, 요소 다시 찾는 중...")
                    
                    print("[Error] 요소를 찾을 수 없습니다.")
                    
                    element = WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#summary_pfm_total > td:nth-child(2)")))
                    
                    the_num = element.text
                    print(the_num)

                # 서비스 계정 키 파일 경로
                credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                # gspread 클라이언트 초기화
                client = gspread.service_account(filename=credential_file)

                # Google 시트 열기
                spreadsheet = client.open_by_url(sheet_url)

                # 첫 번째 시트 선택
                sheet = spreadsheet.worksheet(sheet_name)
                todayy = today.strftime("%Y-%m-%d")
                column_values = sheet.col_values(1)
                for idx, cell_value in enumerate(column_values, start=1):  # start=1로 설정하여 행 번호를 1부터 시작
                    if cell_value == todayy:
                        print(cell_value)
                        print(gspread.utils.rowcol_to_a1(idx, 1))
                        cell_addr = gspread.utils.rowcol_to_a1(idx, 1)
                        # return f"{gspread.utils.rowcol_to_a1(idx, 1)}"  # 셀 주소 반환
                    
                (start_row, start_col) = gspread.utils.a1_to_rowcol(cell_addr)

                # Google 시트에 데이터 쓰기
                numeric_value = int(the_num.replace(',', ''))
                range_to_write = f'C{start_row-ii}'
                sheet.update([[numeric_value]], range_to_write)

            driver.close()
            driver.switch_to.window(driver.window_handles[0]) #팝업으로 제어 변경
            driver.close()

            # 완료 로그
            self.logBox.append(f"{self.etc_group_box.title()}-{self.visitors.text()}-{brand}<br><span style='color:blue;'>완료</span>")

            # except Exception as e:
            #     print(f"Error occurred: {e}")
            #     if driver:
            #         try:
            #             for handle in driver.window_handles:
            #                 driver.switch_to.window(handle)  # 각 창으로 전환
            #                 driver.close()  # 현재 창 닫기
            #         except Exception as close_error:
            #             print(f"Error closing driver: {close_error}")

            #     # 실패 로그
            #     self.logBox.append(f"{self.etc_group_box.title()}-{self.visitors.text()}-{brand}<br><span style='color:red;'>실패</span>")

        #카페24 노마셀
        if self.know_visitors.isChecked() == True:

            url_cafe24 = "https://eclogin.cafe24.com/Shop/" 

            cafe24_id_knowmycell = self.login_info("CAFE_KNOW_ID")
            cafe24_pw_knowmycell = self.login_info("CAFE_KNOW_PW")
            brand = "노마셀"

            sheet_knowR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=567505346'
            sheet_knowD = "노마셀D"

            visitors(url_cafe24, cafe24_id_knowmycell, cafe24_pw_knowmycell, sheet_knowR_url, sheet_knowD, brand)

        #카페24 제니크
        if self.zq_visitors.isChecked() == True:

            url_cafe24 = "https://eclogin.cafe24.com/Shop/" 

            cafe24_id_zq = self.login_info("CAFE_ZQ_ID")
            cafe24_pw_zq = self.login_info("CAFE_ZQ_PW")
            brand = "제니크"

            sheet_zqR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=567505346'
            sheet_zqD = "제니크D"

            visitors(url_cafe24, cafe24_id_zq, cafe24_pw_zq, sheet_zqR_url, sheet_zqD, brand)

# 신규 가입자
         
        def new_member(url, id, pw, sheet_url, sheet_name, brand):

            driver = None
            try:

                # 크롬 On
                chromedriver_path = chromedriver_autoinstaller.install()
                chrome_options = webdriver.ChromeOptions()
                chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
                chrome_options.add_argument("--start-maximized") #최대 크기로 시작
                chrome_options.add_experimental_option('detach', True)

                user_data = self.chrome_path_folder.text()
                chrome_options.add_argument(f"user-data-dir={user_data}")
                chrome_options.add_argument("--profile-directory=Profile 1")
            
                user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
                headers = {'user-agent' : user_agent}

                driver = webdriver.Chrome(
                    service=Service(chromedriver_path),
                    options=chrome_options
                )

                driver.get(url)

                ##################################### 로그인
                ##################################### 로그인
                ##################################### 로그인
                ##################################### 로그인

                # ID
                input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#mall_id")))
                input_field.click()
                time.sleep(1)
                input_field.send_keys(Keys.CONTROL + "a")
                input_field.send_keys(Keys.BACKSPACE)
                driver.find_element(By.CSS_SELECTOR, "#mall_id").send_keys(id)

                # PW
                input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#userpasswd")))
                input_field.click()
                input_field.send_keys(Keys.CONTROL + "a")
                input_field.send_keys(Keys.BACKSPACE)
                driver.find_element(By.CSS_SELECTOR, "#userpasswd").send_keys(pw)

                # 로그인클릭
                driver.find_element(By.CSS_SELECTOR,'#frm_user > div > div.mButton > button').click()

                #비밀번호변경안내
                try: WebDriverWait(driver, 5).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#iptBtnEm")))).click() 
                except: pass

                try:
                    time.sleep(3)
                    popup = driver.find_element(By.XPATH, '//*[contains(text(), "오늘 하루 보지 않기")]')
                    popup.click()

                except: pass

                #화면로딩대기
                WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "오늘의 할 일")]')))
                ###################################### 신규회원
                ###################################### 신규회원
                ###################################### 신규회원
                ###################################### 신규회원

                driver.find_element(By.CSS_SELECTOR, "#mCSB_2_container > ul:nth-child(1) > li:nth-child(9)").click() #통계 클릭
                WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, "//*[contains(text(), '접속통계')]"))).click() #접속통계클릭 #접속통계클릭

                #새 창 대기
                current_window_handle = driver.current_window_handle

                new_window_handle = None
                while not new_window_handle:
                    for handle in driver.window_handles:
                        if handle != current_window_handle:
                            new_window_handle = handle
                            break


                #팝업으로 제어 변경
                driver.switch_to.window(driver.window_handles[1]) 

                WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_vis")))).click() #방문자분석
                WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_vis_nmb")))).click() #신규회원수
                WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(3) > a:nth-child(2) > img")))).click()

                # 서비스 계정 키 파일 경로
                credential_file = 'triple-nectar-412808-da4dac0cc16e.json'
                # gspread 클라이언트 초기화
                client = gspread.service_account(filename=credential_file)
                # Google 시트 열기
                spreadsheet = client.open_by_url(sheet_url)
                # 첫 번째 시트 선택
                sheet = spreadsheet.worksheet(sheet_name)

                # 2번째 열의 모든 값을 가져옵니다.
                col_values = sheet.col_values(2)

                # 마지막 행 번호 찾기
                last_row = len(col_values)

                for ii in range(1, target_days_input+1):

                    # 달력클릭
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(2) > a:nth-child(3)")))).click()

                    days_vis = datetime.timedelta(days=ii)
                    before_day_vis = today-days_vis
                    

                    #시작
                    #년도 선택
                    before_year = (before_day_vis).strftime("%Y")
                    select_element = driver.find_element(By.CSS_SELECTOR, '#yearSB_1')
                    select = Select(select_element)
                    select.select_by_value(before_year)

                    #달 선택
                    before_month = str(int((before_day_vis).strftime("%m")))
                    select_element = driver.find_element(By.CSS_SELECTOR, '#monthSB_1')
                    select = Select(select_element)
                    select.select_by_value(before_month)

                    #일 선택
                    before_day1 = str(int((before_day_vis).strftime("%d")))
                    for i in range(1, 43):
                        try:
                            element = driver.find_element(By.ID, f'li_{i}')
                            if element.text == before_day1:
                                element.click()
                                print("before_day1 clicked")
                                break

                        except:
                            print(f'li_{i} not found')


                    #끝
                    #년도 선택
                    select_element = driver.find_element(By.CSS_SELECTOR, '#yearSB_2')
                    select = Select(select_element)
                    select.select_by_value(before_year)

                    #달 선택
                    select_element = driver.find_element(By.CSS_SELECTOR, '#monthSB_2')
                    select = Select(select_element)
                    select.select_by_value(before_month)

                    #일 선택
                    for i in range(1, 43):
                        try:
                            element = driver.find_element(By.ID, f'le_{i}')
                            if element.text == before_day1:
                                element.click()
                                print("before_day1 clicked")
                                break

                        except:
                            print(f'le_{i} not found')

                    # 조회
                    try:
                        element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(2) > a:nth-child(4) > img")))
                        
                        element.click()

                    except:
                        print("재시도")

                        element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(2) > a:nth-child(4) > img")))

                        time.sleep(1)
                        
                        element.click()

                    
                    total_nmb = driver.find_element(By.CSS_SELECTOR, "#tbl_data > tbody > tr").text
                    total_nmb_list = total_nmb.split()
                    print(total_nmb_list)

                    
                    todayy = total_nmb_list[0]+total_nmb_list[1]
                    print(todayy)
                    column_values = sheet.col_values(2)
                    for idx, cell_value in enumerate(column_values, start=1):  # start=1로 설정하여 행 번호를 1부터 시작
                        if cell_value == todayy:
                            print("cell_value == todayy")
                            print(cell_value)
                            print(gspread.utils.rowcol_to_a1(idx, 2))
                            cell_addr = gspread.utils.rowcol_to_a1(idx, 2)
                            # return f"{gspread.utils.rowcol_to_a1(idx, 1)}"  # 셀 주소 반환
                            break
                        
                        else:
                            print("col_values = sheet.col_values(2)")

                            cell_addr = gspread.utils.rowcol_to_a1(last_row+target_days_input+1, 2)
                            print(cell_addr)
                            break

                        
                    (start_row, start_col) = gspread.utils.a1_to_rowcol(cell_addr)

                    # Google 시트에 데이터 쓰기
                    numeric_value = int(total_nmb_list[2].replace(',', ''))
                    range_to_write = f'B{start_row-ii}:D{start_row-ii}'
                    sheet.update([[total_nmb_list[0]+total_nmb_list[1],numeric_value,total_nmb_list[3]]], range_to_write)

                driver.close()
                driver.switch_to.window(driver.window_handles[0]) #팝업으로 제어 변경
                driver.close()

                # 완료 로그
                self.logBox.append(f"{self.etc_group_box.title()}-{self.newMemb.text()}-{brand}<br><span style='color:blue;'>완료</span>")

            except Exception as e:
                print(f"Error occurred: {e}")
                if driver:
                    try:
                        for handle in driver.window_handles:
                            driver.switch_to.window(handle)  # 각 창으로 전환
                            driver.close()  # 현재 창 닫기
                    except Exception as close_error:
                        print(f"Error closing driver: {close_error}")

                # 실패 로그
                self.logBox.append(f"{self.etc_group_box.title()}-{self.newMemb.text()}-{brand}<br><span style='color:red;'>실패</span>")


        sheet_url = "https://docs.google.com/spreadsheets/d/1zZ4jOfvMavEEuXtvGu0FCnZBAltbS6KQUjJp-aS9C0Q/edit?gid=1144250812#gid=1144250812"

        #카페24 노마셀
        if self.know_newMemb.isChecked() == True:

            url_cafe24 = "https://eclogin.cafe24.com/Shop/" 

            cafe24_id_knowmycell = self.login_info("CAFE_KNOW_ID")
            cafe24_pw_knowmycell = self.login_info("CAFE_KNOW_PW")
            brand = "노마셀"
            sheet_name = "노마셀 신규회원수"

            new_member(url_cafe24, cafe24_id_knowmycell, cafe24_pw_knowmycell, sheet_url, sheet_name, brand)

        #카페24 제니크
        if self.zq_newMemb.isChecked() == True:

            url_cafe24 = "https://eclogin.cafe24.com/Shop/" 

            cafe24_id_zq = self.login_info("CAFE_ZQ_ID")
            cafe24_pw_zq = self.login_info("CAFE_ZQ_PW")
            brand = "제니크"
            sheet_name = "제니크 신규회원수"

            new_member(url_cafe24, cafe24_id_zq, cafe24_pw_zq, sheet_url, sheet_name, brand)


        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("추출이 완료되었습니다")
        msg.setWindowTitle("알림")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setWindowFlags(msg.windowFlags() | Qt.WindowStaysOnTopHint)
        msg.exec_()
        return   

    def saveText(self):
        text = self.path_folder.text()
        text1 = self.chrome_path_folder.text()
        text2 = self.edge_path_folder.text()
        with open('saved_text.txt', 'w') as file:
            file.write(text)
            file.write("\n")
            file.write(text1)
            file.write("\n")
            file.write(text2)
        QMessageBox.information(self,'알림','저장되었습니다.')

        with open('checkbox_state.txt', 'w') as file:
            file.write(f"{self.CMlabs_salesCafe24.isChecked()}\n")
            file.write(f"{self.know_salesCafe24.isChecked()}\n")
            file.write(f"{self.zq_salesCafe24.isChecked()}\n")

            file.write(f"{self.know_salesCoup.isChecked()}\n")

            file.write(f"{self.know_salesNaver.isChecked()}\n")

            file.write(f"{self.know_advtCoup.isChecked()}\n")

            file.write(f"{self.know_advtNaver.isChecked()}\n")
            file.write(f"{self.zq_advtNaver.isChecked()}\n")

            file.write(f"{self.know_advtGFA.isChecked()}\n")
            file.write(f"{self.zq_advtGFA.isChecked()}\n")

            file.write(f"{self.know_advtPC.isChecked()}\n")

            file.write(f"{self.know_advtGgle.isChecked()}\n")

            file.write(f"{self.CMlabs_advtMeta.isChecked()}\n")
            file.write(f"{self.know_advtMeta.isChecked()}\n")
            file.write(f"{self.zq_advtMeta.isChecked()}\n")

            file.write(f"{self.CMlabs_advtTiktok.isChecked()}\n")
            file.write(f"{self.know_advtTiktok.isChecked()}\n")
            file.write(f"{self.zq_advtTiktok.isChecked()}\n")

            file.write(f"{self.know_visitors.isChecked()}\n")
            file.write(f"{self.zq_visitors.isChecked()}\n")

            file.write(f"{self.know_newMemb.isChecked()}\n")
            file.write(f"{self.zq_newMemb.isChecked()}\n")

    def loadCheckboxState(self):
        try:
            with open('checkbox_state.txt', 'r') as file:
                states = file.readlines()
                self.CMlabs_salesCafe24.setChecked(states[0].strip() == 'True')
                self.know_salesCafe24.setChecked(states[1].strip() == 'True')
                self.zq_salesCafe24.setChecked(states[2].strip() == 'True')

                self.know_salesCoup.setChecked(states[3].strip() == 'True')

                self.know_salesNaver.setChecked(states[4].strip() == 'True')

                self.know_advtCoup.setChecked(states[5].strip() == 'True')

                self.know_advtNaver.setChecked(states[6].strip() == 'True')
                self.zq_advtNaver.setChecked(states[7].strip() == 'True')

                self.know_advtGFA.setChecked(states[8].strip() == 'True')
                self.zq_advtGFA.setChecked(states[9].strip() == 'True')

                self.know_advtPC.setChecked(states[10].strip() == 'True')

                self.know_advtGgle.setChecked(states[11].strip() == 'True')

                self.CMlabs_advtMeta.setChecked(states[12].strip() == 'True')
                self.know_advtMeta.setChecked(states[13].strip() == 'True')
                self.zq_advtMeta.setChecked(states[14].strip() == 'True')

                self.CMlabs_advtTiktok.setChecked(states[15].strip() == 'True')
                self.know_advtTiktok.setChecked(states[16].strip() == 'True')
                self.zq_advtTiktok.setChecked(states[17].strip() == 'True')

                self.know_visitors.setChecked(states[18].strip() == 'True')
                self.zq_visitors.setChecked(states[19].strip() == 'True')

                self.know_newMemb.setChecked(states[20].strip() == 'True')
                self.zq_newMemb.setChecked(states[21].strip() == 'True')
                # 나머지 체크박스도 동일하게 불러옵니다.
        except FileNotFoundError:
            pass

    def loadText(self):
            try:
                with open('saved_text.txt', 'r') as f:
                    saved_text = f.read()
                    texts = saved_text.split("\n")

                    print(texts)

                    self.path_folder.setText(texts[0])
                    self.chrome_path_folder.setText(texts[1])
                    self.edge_path_folder.setText(texts[2])

                    
            except FileNotFoundError:
                pass

    def login_info(self, target_word):
        try:
            with open('login_info.txt', 'r', encoding='utf-8') as f:
                lines = f.readlines()  # 파일의 모든 줄을 읽어 리스트로 저장

            # 모든 줄을 순회하면서 target_word 찾기
            for i, line in enumerate(lines):
                if target_word in line:  # 현재 줄에 target_word가 포함되어 있는지 확인
                    if i + 1 < len(lines):  # 다음 줄이 존재하는지 확인
                        print(lines[i + 1].strip())  # 다음 줄의 내용을 프린트 (공백 제거)
                        return(lines[i + 1].strip())
        except FileNotFoundError: print("cannot find login information.")
        
    def folderopen(self):
        fname = QFileDialog.getExistingDirectory(self,'폴더선택','')
        self.path_folder.setText(fname)
    
    def chromefolderopen(self):
        fname = QFileDialog.getExistingDirectory(self,'폴더선택','')
        self.chrome_path_folder.setText(fname)

    def edgefolderopen(self):
        fname = QFileDialog.getExistingDirectory(self,'폴더선택','')
        self.edge_path_folder.setText(fname)

    def my_exception_hook(exctype, value, traceback):
        # Print the error and traceback
        print(exctype, value, traceback)
        # Call the normal Exception hook after
        sys._excepthook(exctype, value, traceback)
        # sys.exit(1)

    # Back up the reference to the exceptionhook
    sys._excepthook = sys.excepthook

    # Set the exception hook to our wrapping function
    sys.excepthook = my_exception_hook

if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = Rawdata_extractor()
    win.show()
    sys.exit(app.exec_())
